import io
import json
import os
import time
from datetime import date, datetime, timezone
from typing import Dict, List, Optional

import pandas as pd
import streamlit as st
import bcrypt
from streamlit_searchbox import st_searchbox
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from firebase_admin import credentials, firestore, storage
import firebase_admin

COD_CLIENTE_IMPAKTO = 208831
FIRESTORE_COLLECTION = "pedido_itens"
FIRESTORE_PRODUCTS_COLLECTION = "produtos"
FIRESTORE_SETORES_COLLECTION = "setores"
FIRESTORE_PDF_BUCKET = "material-basico"
FIRESTORE_AUDIT_COLLECTION = "audit_trail"
FIRESTORE_USERS_COLLECTION = "users"
FIRESTORE_PRODUCT_HISTORY_COLLECTION = "historico_produtos"
FIRESTORE_SETORES_HISTORY_COLLECTION = "historico_setores"
COMPANY_NAME = "IMPAKTO"
LOGO_CANDIDATE_FILES = [
    "logo.png",
    "logo.jpg",
    "logo.jpeg",
    "logo.webp",
    "assets/logo.png",
    "assets/logo.jpg",
    "assets/logo.jpeg",
    "assets/logo.webp",
]
AVATAR_CANDIDATE_FILES = [
    "assets/avatar.png",
    "assets/avatar.jpg",
    "assets/avatar.jpeg",
    "assets/user.png",
    "assets/user.jpg",
    "avatar.png",
    "avatar.jpg",
    "user.png",
    "user.jpg",
]
ENABLE_PDF_STORAGE = os.getenv("ENABLE_PDF_STORAGE", "false").lower() == "true"  # Disabled by default
APP_PASSWORD = os.getenv("APP_PASSWORD", "admin123")  # Default for testing
SESSION_DEFAULTS = {
    "excel_data": None,
    "excel_source": "",
    "excel_bytes": b"",
    "cart": [],
    "selected_setor": None,
    "firestore_client": None,
    "edit_client": None,
    "edit_indices": [],
    "authenticated": False,
    "current_user": None,
    "layout_mode_new": False,
    "product_update_changes": pd.DataFrame(),
    "product_update_df": pd.DataFrame(),
    "product_update_source": "manual",
    "_admin_checked": False,
    "_firestore_synced": False,
}
SHEET_CONFIG = {
    "Produtos": None,
    "Setor": None,
    "Aguardando Aprovação": [
        "CòdClienteImpakto",
        "CódProImpakto",
        "Item",
        "Qtde",
        "$ Unitário",
        "$ Total",
        "Unidade",
        "Setor",
        "SETOR2",
    ],
    "Pedido": [
        "CòdClienteImpakto",
        "CódProImpakto",
        "Item",
        "Qtde",
        "$ Unitário",
        "$ Total",
        "Unidade",
        "Setor",
        "SETOR2",
    ],
}


def init_session_state() -> None:
    for key, value in SESSION_DEFAULTS.items():
        st.session_state.setdefault(key, value)

    # Run once per session — not on every Streamlit rerun
    if not st.session_state._admin_checked:
        ensure_admin_user()
        st.session_state._admin_checked = True

    if not st.session_state._firestore_synced and firestore_enabled():
        st.session_state._firestore_synced = True  # mark before sync so crashes don't retry infinitely
        sync_firestore_to_session()


def check_authentication() -> bool:
    """Simple password authentication."""
    if st.session_state.authenticated:
        return True
    
    if not firestore_enabled():
        st.session_state.authenticated = True
        st.session_state.current_user = "Offline"
        return True
    
    return False


def render_login() -> bool:
    """Render login page with Firestore user validation."""
    st.set_page_config(page_title="Sistema de Gestão - Login", layout="centered")
    st.markdown(
        """
        <style>
        .stApp {
            background:
                radial-gradient(920px 420px at 50% -30%, rgba(255, 255, 255, 0.2), transparent 70%),
                linear-gradient(180deg, #5b59cc 0%, #4e4db2 100%);
        }

        [data-testid="stAppViewContainer"] > .main .block-container {
            max-width: 720px;
            padding-top: 3rem;
        }

        [data-testid="stForm"],
        [data-testid="stVerticalBlock"] > [data-testid="element-container"] {
            border-radius: 16px;
        }

        [data-testid="stTextInput"] input {
            border-radius: 12px !important;
            border: 1px solid #c7d2fe !important;
            min-height: 44px;
            background: #ffffff !important;
        }

        [data-testid="stTextInput"] input:focus {
            border-color: #818cf8 !important;
            box-shadow: 0 0 0 1px #818cf8 !important;
        }

        .stButton > button,
        [data-testid="baseButton-primary"] {
            border-radius: 12px !important;
            min-height: 44px;
            border: 1px solid #4e4db2 !important;
            background: linear-gradient(90deg, #5b59cc 0%, #4e4db2 100%) !important;
            color: #ffffff !important;
            font-weight: 600;
        }

        [data-testid="stAlertContainer"] {
            border-radius: 12px;
        }

        .login-title {
            margin: 0;
            color: #ffffff;
            text-align: center;
            font-size: 2rem;
            font-weight: 700;
        }

        .login-logo {
            display: flex;
            justify-content: center;
            margin-bottom: 0.65rem;
        }

        .login-subtitle {
            margin: 0.3rem 0 0.9rem 0;
            color: rgba(255, 255, 255, 0.92);
            text-align: center;
            font-size: 0.96rem;
        }

        .login-card {
            background: rgba(255, 255, 255, 0.82);
            border: 1px solid rgba(255, 255, 255, 0.65);
            border-radius: 20px;
            box-shadow: 0 24px 54px rgba(2, 46, 86, 0.2);
            backdrop-filter: blur(12px);
            padding: 1rem 1rem 0.9rem 1rem;
        }

        .login-card hr {
            border-color: rgba(148, 163, 184, 0.3);
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    logo_path = get_company_logo_path()
    if logo_path:
        st.markdown('<div class="login-logo">', unsafe_allow_html=True)
        st.image(logo_path, width=220)
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<p class="login-title">🔐 Login</p>', unsafe_allow_html=True)
    st.markdown('<p class="login-subtitle">Sistema de Gestão de Pedidos</p>', unsafe_allow_html=True)
    st.markdown('<div class="login-card">', unsafe_allow_html=True)
    st.divider()
    
    col1, col2 = st.columns([1, 1])
    with col1:
        username = st.text_input("👤 Usuário")
    with col2:
        password = st.text_input("🔑 Senha", type="password")
    
    if st.button("Entrar", use_container_width=True):
        if validate_user_credentials(username, password):
            st.session_state.authenticated = True
            st.session_state.current_user = username if username else "Usuario"
            st.success(f"✅ Bem-vindo, {st.session_state.current_user}!")
            st.rerun()
        else:
            st.error("❌ Usuário ou senha incorretos!")
    
    st.divider()
    st.markdown("</div>", unsafe_allow_html=True)


def log_audit(action: str, details: Dict) -> None:
    """Log audit trail to Firestore."""
    if not firestore_enabled():
        return
    
    try:
        db = get_firestore_client()
        audit_log = {
            "action": action,
            "user": st.session_state.get("current_user", "unknown"),
            "timestamp": firestore.SERVER_TIMESTAMP,
            "details": details,
        }
        db.collection(FIRESTORE_AUDIT_COLLECTION).add(audit_log)
    except Exception as exc:
        st.warning(f"⚠️ Erro ao registrar auditoria: {exc}")


def compare_orders(original: Dict, modified: Dict) -> Dict:
    """Compare original and modified orders to get changes."""
    changes = {}
    for key in ["Qtde", "Item", "$ Unitário", "$ Total"]:
        orig_val = original.get(key)
        mod_val = modified.get(key)
        if orig_val != mod_val:
            changes[key] = {"de": orig_val, "para": mod_val}
    return changes


# ============================================================================
# USER MANAGEMENT FUNCTIONS
# ============================================================================


def hash_password(password: str) -> str:
    """Hash a password using bcrypt."""
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')


def verify_password(password: str, hashed: str) -> bool:
    """Verify a password against its hash."""
    try:
        return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))
    except Exception:
        return False


def get_user_info(username: str) -> Dict:
    """Get user information from Firestore."""
    if not firestore_enabled() or not username:
        return {"role": "user", "assigned_setores": []}
    
    try:
        db = get_firestore_client()
        doc = db.collection(FIRESTORE_USERS_COLLECTION).document(username).get()
        if doc.exists:
            user_data = doc.to_dict()
            return {
                "role": user_data.get("role", "user"),
                "assigned_setores": user_data.get("assigned_setores", [])
            }
    except:
        pass
    
    return {"role": "user", "assigned_setores": []}


def can_edit_catalog() -> bool:
    current_user = st.session_state.get("current_user")
    if current_user == "admin":
        return True
    user_info = get_user_info(current_user)
    return user_info.get("role") in {"admin", "catalog_editor"}


def filter_setores_for_user(setores_df: pd.DataFrame, user_info: Dict) -> pd.DataFrame:
    """Filter setores based on user role and assignments."""
    # Admin and user see all setores
    if user_info["role"] != "supervisora":
        return setores_df
    
    # Supervisora - filter by assigned setores
    if not user_info.get("assigned_setores"):
        return setores_df
    
    # Create label column if not exists (needed for filtering)
    if "label" not in setores_df.columns and "CódUnidade" in setores_df.columns and "items__description" in setores_df.columns:
        setores_df = setores_df.copy()
        setores_df["codigo"] = setores_df["CódUnidade"].apply(parse_int)
        setores_df["descricao"] = setores_df["items__description"].astype(str)
        setores_df["label"] = setores_df.apply(
            lambda row: f"{row['codigo']} - {row['descricao']}", axis=1
        )
    
    # Filter only assigned setores
    filtered = setores_df[setores_df["label"].isin(user_info["assigned_setores"])]
    return filtered


def get_users_list() -> List[Dict]:
    """Fetch all users from Firestore."""
    if not firestore_enabled():
        return []
    
    try:
        db = get_firestore_client()
        users = []
        for doc in db.collection(FIRESTORE_USERS_COLLECTION).stream():
            user_data = doc.to_dict()
            user_data["username"] = doc.id
            users.append(user_data)
        return users
    except Exception as e:
        st.warning(f"⚠️ Erro ao carregar usuários: {e}")
        return []


def validate_user_credentials(username: str, password: str) -> bool:
    """Validate user credentials against Firestore users collection."""
    if not firestore_enabled():
        # Offline mode: allow with default password
        return password == APP_PASSWORD
    
    if not username or not password:
        return False
    
    try:
        db = get_firestore_client()
        doc = db.collection(FIRESTORE_USERS_COLLECTION).document(username).get()
        
        if not doc.exists:
            return False
        
        user_data = doc.to_dict()
        password_hash = user_data.get("password_hash", "")
        
        return verify_password(password, password_hash)
    except Exception as e:
        st.warning(f"⚠️ Erro ao validar credenciais: {e}")
        return False


def create_user(username: str, password: str, role: str = "user", assigned_setores: List[str] = None) -> bool:
    """Create a new user in Firestore."""
    if not firestore_enabled():
        st.error("❌ Firestore não disponível para criar usuários.")
        return False
    
    if not username or not password:
        st.error("❌ Usuário e senha são obrigatórios.")
        return False
    
    try:
        db = get_firestore_client()
        
        # Check if user already exists
        doc = db.collection(FIRESTORE_USERS_COLLECTION).document(username).get()
        if doc.exists:
            st.error(f"❌ Usuário '{username}' já existe.")
            return False
        
        # Create user
        user_data = {
            "password_hash": hash_password(password),
            "role": role,
            "created_at": firestore.SERVER_TIMESTAMP,
            "updated_at": firestore.SERVER_TIMESTAMP,
        }
        
        # Add assigned setores for supervisora role
        if role == "supervisora" and assigned_setores:
            user_data["assigned_setores"] = assigned_setores
        
        db.collection(FIRESTORE_USERS_COLLECTION).document(username).set(user_data)
        st.success(f"✅ Usuário '{username}' criado com sucesso!")
        log_audit("create_user", {"username": username, "role": role, "assigned_setores": assigned_setores})
        return True
    except Exception as e:
        st.error(f"❌ Erro ao criar usuário: {e}")
        return False


def delete_user(username: str) -> bool:
    """Delete a user from Firestore."""
    if not firestore_enabled():
        st.error("❌ Firestore não disponível.")
        return False
    
    try:
        db = get_firestore_client()
        doc_ref = db.collection(FIRESTORE_USERS_COLLECTION).document(username)
        doc = doc_ref.get()
        if not doc.exists:
            message = f"❌ Usuário '{username}' não encontrado."
            st.session_state["user_action_error"] = message
            st.error(message)
            return False
        doc_ref.delete()
        log_audit("delete_user", {"username": username})
        return True
    except Exception as e:
        message = f"❌ Erro ao deletar usuário: {e}"
        st.session_state["user_action_error"] = message
        st.error(message)
        return False


def update_user_password(username: str, new_password: str) -> bool:
    """Update a user's password."""
    if not firestore_enabled():
        st.error("❌ Firestore não disponível.")
        return False
    
    if not new_password:
        message = "❌ Nova senha é obrigatória."
        st.session_state["user_action_error"] = message
        st.error(message)
        return False
    
    try:
        db = get_firestore_client()
        db.collection(FIRESTORE_USERS_COLLECTION).document(username).update({
            "password_hash": hash_password(new_password),
            "updated_at": firestore.SERVER_TIMESTAMP,
        })
        st.success(f"✅ Senha do usuário '{username}' atualizada com sucesso!")
        log_audit("update_user_password", {"username": username})
        return True
    except Exception as e:
        message = f"❌ Erro ao atualizar senha: {e}"
        st.session_state["user_action_error"] = message
        st.error(message)
        return False


def ensure_admin_user():
    """Ensure admin user exists in Firestore on first run."""
    if not firestore_enabled():
        return
    
    try:
        db = get_firestore_client()
        doc = db.collection(FIRESTORE_USERS_COLLECTION).document("admin").get()
        
        if not doc.exists:
            admin_data = {
                "password_hash": hash_password("admin123"),
                "role": "admin",
                "created_at": firestore.SERVER_TIMESTAMP,
                "updated_at": firestore.SERVER_TIMESTAMP,
            }
            db.collection(FIRESTORE_USERS_COLLECTION).document("admin").set(admin_data)
    except Exception as e:
        st.warning(f"⚠️ Erro ao garantir usuário admin: {e}")


def ensure_columns(df: pd.DataFrame, columns: List[str]) -> pd.DataFrame:
    if df is None:
        return pd.DataFrame(columns=columns)
    for col in columns:
        if col not in df.columns:
            df[col] = None
    return df[columns].copy()


def load_excel_data(bytes_data: bytes) -> Dict[str, pd.DataFrame]:
    buffer = io.BytesIO(bytes_data)
    workbook = pd.ExcelFile(buffer)
    sheets: Dict[str, pd.DataFrame] = {}

    for sheet_name, columns in SHEET_CONFIG.items():
        if sheet_name in workbook.sheet_names:
            df = workbook.parse(sheet_name)
        else:
            df = pd.DataFrame(columns=columns or [])

        if columns:
            df = ensure_columns(df, columns)

        sheets[sheet_name] = df

    for sheet_name in ("Produtos", "Setor"):
        if sheet_name not in sheets:
            sheets[sheet_name] = pd.DataFrame()

    return sheets


def set_excel_data(bytes_data: bytes, source_label: str) -> None:
    try:
        excel_data = load_excel_data(bytes_data)
    except Exception as exc:  # pragma: no cover - streamlit feedback
        st.error(f"Falha ao carregar planilha: {exc}")
        return

    st.session_state.excel_data = excel_data
    st.session_state.excel_source = source_label
    st.session_state.excel_bytes = bytes_data
    st.session_state.cart = []
    st.session_state.selected_setor = None
    st.toast(f"Planilha carregada de {source_label}", icon="✅")

def get_firestore_client():
    if st.session_state.get("firestore_client"):
        return st.session_state.firestore_client

    service_json = os.getenv("FIREBASE_SERVICE_ACCOUNT_JSON")
    if not service_json:
        return None

    try:
        cred_dict = json.loads(service_json)
    except json.JSONDecodeError:
        st.error("FIREBASE_SERVICE_ACCOUNT_JSON invalido.")
        return None

    if not firebase_admin._apps:
        cred = credentials.Certificate(cred_dict)
        firebase_admin.initialize_app(cred, {
            'storageBucket': FIRESTORE_PDF_BUCKET
        })

    client = firestore.client()
    st.session_state.firestore_client = client
    return client


def firestore_enabled() -> bool:
    return get_firestore_client() is not None


def fetch_firestore_rows(status: Optional[str] = None) -> pd.DataFrame:
    db = get_firestore_client()
    if not db:
        return pd.DataFrame()

    query = db.collection(FIRESTORE_COLLECTION)
    if status:
        query = query.where("status", "==", status)

    rows = []
    for attempt in range(3):
        try:
            rows = []
            for doc in query.stream():
                data = doc.to_dict()
                data["__doc_id"] = doc.id
                rows.append(data)
            break
        except Exception as e:
            if "429" in str(e) or "Quota" in str(e) or "ResourceExhausted" in str(e):
                if attempt < 2:
                    time.sleep(2 ** attempt)
                    continue
                st.warning("⚠️ Cota do Firestore excedida. Tente novamente em alguns instantes.")
                return pd.DataFrame()
            raise

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    df = ensure_columns(df, SHEET_CONFIG["Aguardando Aprovação"])
    df["__doc_id"] = pd.Series([row["__doc_id"] for row in rows])
    return df


def fetch_firestore_rows_raw(status: Optional[str] = None) -> pd.DataFrame:
    db = get_firestore_client()
    if not db:
        return pd.DataFrame()

    query = db.collection(FIRESTORE_COLLECTION)
    if status:
        query = query.where("status", "==", status)

    rows = []
    for doc in query.stream():
        data = doc.to_dict()
        data["__doc_id"] = doc.id
        rows.append(data)

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    for col in SHEET_CONFIG["Aguardando Aprovação"]:
        if col not in df.columns:
            df[col] = None
    return df


def sync_firestore_to_session() -> None:
    if not firestore_enabled():
        return

    aguardando = fetch_firestore_rows("aguardando")
    pedido = fetch_firestore_rows("pedido")

    if st.session_state.excel_data is None:
        st.session_state.excel_data = {}

    st.session_state.excel_data["Aguardando Aprovação"] = aguardando
    st.session_state.excel_data["Pedido"] = pedido
    sync_firestore_reference_data()


def upload_pdf_to_storage(pdf_buffer: io.BytesIO, client_name: str) -> Optional[str]:
    """Upload PDF to Firestore Storage and return the path."""
    if not ENABLE_PDF_STORAGE:
        return None
    
    try:
        # Get Firebase Admin SDK instance
        if not firebase_admin._apps:
            get_firestore_client()  # Initialize if needed
        
        bucket = storage.bucket(FIRESTORE_PDF_BUCKET)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        pdf_name = f"previews/{client_name.replace(' ', '_')}_PREVIA_{timestamp}.pdf"
        blob = bucket.blob(pdf_name)
        
        # Upload PDF
        blob.upload_from_string(
            pdf_buffer.getvalue(), 
            content_type="application/pdf"
        )
        st.success(f"✅ PDF salvo em Storage: {pdf_name}")
        return pdf_name
    except Exception as exc:
        # If bucket doesn't exist or upload fails, just log it and continue
        # The order is still saved in Firestore which is the important part
        st.warning(f"⚠️ Não foi possível salvar PDF em Storage (storage não configurado), mas o pedido foi salvo normalmente.")
        return None


def get_next_order_number() -> int:
    """Get the next order number using Firestore transaction."""
    db = get_firestore_client()
    if not db:
        return 1
    
    counter_ref = db.collection("counters").document("order_counter")
    
    @firestore.transactional
    def increment_counter(transaction, ref):
        snapshot = ref.get(transaction=transaction)
        if snapshot.exists:
            current = snapshot.to_dict().get("value", 0)
        else:
            current = 0
        
        next_value = current + 1
        transaction.set(ref, {"value": next_value})
        return next_value
    
    try:
        transaction = db.transaction()
        next_number = increment_counter(transaction, counter_ref)
        return next_number
    except Exception as e:
        st.warning(f"⚠️ Erro ao gerar número de pedido: {e}. Usando timestamp.")
        # Fallback to timestamp-based number
        return int(datetime.now().timestamp())


def generate_pending_order_id() -> str:
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S%f")
    return f"pend_{timestamp}_{os.urandom(3).hex()}"


def acquire_edit_lock(client_name: str, username: str) -> bool:
    """Try to acquire an edit lock for a client's order. Returns True if successful."""
    db = get_firestore_client()
    if not db:
        return True  # Allow editing if Firestore is not available
    
    try:
        lock_ref = db.collection("edit_locks").document(client_name)
        
        @firestore.transactional
        def try_lock(transaction, ref):
            snapshot = ref.get(transaction=transaction)
            
            if snapshot.exists:
                lock_data = snapshot.to_dict()
                locked_by = lock_data.get("locked_by")
                locked_at = lock_data.get("locked_at")
                
                # Check if lock is stale (older than 10 minutes)
                if locked_at and hasattr(locked_at, 'to_datetime'):
                    lock_time = locked_at.to_datetime()
                    if (datetime.now(timezone.utc) - lock_time).total_seconds() > 600:  # 10 minutes
                        # Lock is stale, release it
                        transaction.set(ref, {
                            "locked_by": username,
                            "locked_at": firestore.SERVER_TIMESTAMP,
                        })
                        return True
                
                # Lock exists and is not stale
                if locked_by == username:
                    # Same user, refresh the lock
                    transaction.update(ref, {"locked_at": firestore.SERVER_TIMESTAMP})
                    return True
                else:
                    # Locked by another user
                    return False
            else:
                # No lock exists, create it
                transaction.set(ref, {
                    "locked_by": username,
                    "locked_at": firestore.SERVER_TIMESTAMP,
                })
                return True
        
        transaction = db.transaction()
        return try_lock(transaction, lock_ref)
    except Exception as e:
        st.warning(f"⚠️ Erro ao adquirir lock: {e}")
        return True  # Allow editing on error to avoid blocking users


def release_edit_lock(client_name: str, username: str) -> None:
    """Release the edit lock for a client's order."""
    db = get_firestore_client()
    if not db:
        return
    
    try:
        lock_ref = db.collection("edit_locks").document(client_name)
        snapshot = lock_ref.get()
        
        if snapshot.exists:
            lock_data = snapshot.to_dict()
            if lock_data.get("locked_by") == username:
                lock_ref.delete()
    except Exception as e:
        # Silent failure - locks will expire anyway
        pass


def get_edit_lock_info(client_name: str) -> Optional[Dict]:
    """Get information about who is currently editing a client's order."""
    db = get_firestore_client()
    if not db:
        return None
    
    try:
        lock_ref = db.collection("edit_locks").document(client_name)
        snapshot = lock_ref.get()
        
        if snapshot.exists:
            lock_data = snapshot.to_dict()
            locked_at = lock_data.get("locked_at")
            
            # Check if lock is stale
            if locked_at and hasattr(locked_at, 'to_datetime'):
                lock_time = locked_at.to_datetime()
                if (datetime.now(timezone.utc) - lock_time).total_seconds() > 600:
                    # Lock is stale, can be ignored
                    return None
            
            return lock_data
        return None
    except Exception:
        return None


def save_rows_to_firestore(
    linhas: pd.DataFrame,
    status: str,
    pdf_buffer: Optional[io.BytesIO] = None,
    pending_order_id: Optional[str] = None,
) -> None:
    db = get_firestore_client()
    if not db:
        return

    # Upload PDF if provided
    pdf_path = None
    client_name = None
    if pdf_buffer and st.session_state.selected_setor:
        client_name = st.session_state.selected_setor.get("descricao", "cliente")
        pdf_path = upload_pdf_to_storage(pdf_buffer, client_name)

    order_number = get_next_order_number() if status == "pedido" else None
    if status == "aguardando":
        pending_order_id = pending_order_id or generate_pending_order_id()

    batch = db.batch()
    op_count = 0
    max_ops = 450

    for _, row in linhas.iterrows():
        payload = {
            "CòdClienteImpakto": row["CòdClienteImpakto"],
            "CódProImpakto": row["CódProImpakto"],
            "Item": row["Item"],
            "Qtde": int(row["Qtde"]),
            "$ Unitário": float(row["$ Unitário"]),
            "$ Total": float(row["$ Total"]),
            "Unidade": row["Unidade"],
            "Setor": row["Setor"],
            "SETOR2": row["SETOR2"],
            "status": status,
            "order_number": order_number,
            "pending_order_id": pending_order_id,
            "client_name": client_name,
            "pdf_path": pdf_path,
            "created_at": firestore.SERVER_TIMESTAMP,
            "updated_at": firestore.SERVER_TIMESTAMP,
        }
        doc_ref = db.collection(FIRESTORE_COLLECTION).document()
        batch.set(doc_ref, payload)
        op_count += 1
        if op_count >= max_ops:
            batch.commit()
            batch = db.batch()
            op_count = 0

    if op_count:
        batch.commit()


def fetch_collection_df(collection: str, columns: List[str]) -> pd.DataFrame:
    db = get_firestore_client()
    if not db:
        return pd.DataFrame(columns=columns)

    rows = []
    for attempt in range(3):
        try:
            rows = []
            for doc in db.collection(collection).stream():
                data = doc.to_dict()
                data["__doc_id"] = doc.id
                rows.append(data)
            break
        except Exception as e:
            if "429" in str(e) or "Quota" in str(e) or "ResourceExhausted" in str(e):
                if attempt < 2:
                    time.sleep(2 ** attempt)
                    continue
                st.warning("⚠️ Cota do Firestore excedida. Tente novamente em alguns instantes.")
                return pd.DataFrame(columns=columns)
            raise

    if not rows:
        return pd.DataFrame(columns=columns)

    df = pd.DataFrame(rows)
    df = ensure_columns(df, columns)
    df["__doc_id"] = pd.Series([row["__doc_id"] for row in rows])
    return df


@st.cache_data(show_spinner=False, ttl=600)
def fetch_products_df() -> pd.DataFrame:
    columns = ["productCode", "name", "price"]
    return fetch_collection_df(FIRESTORE_PRODUCTS_COLLECTION, columns)


@st.cache_data(show_spinner=False, ttl=600)
def fetch_setores_df() -> pd.DataFrame:
    columns = ["CódUnidade", "items__description"]
    return fetch_collection_df(FIRESTORE_SETORES_COLLECTION, columns)


def upsert_collection_from_df(collection: str, df: pd.DataFrame, key_field: str) -> None:
    db = get_firestore_client()
    if not db:
        return

    for _, row in df.iterrows():
        doc_id = str(row.get(key_field, "")).strip()
        if not doc_id:
            continue
        payload = row.drop(labels=["__doc_id"], errors="ignore").to_dict()
        db.collection(collection).document(doc_id).set(payload)


def normalize_products_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["productCode", "name", "price"])

    normalized = df.copy()
    for col in ["productCode", "name", "price"]:
        if col not in normalized.columns:
            normalized[col] = None

    normalized = normalized[["productCode", "name", "price"]].copy()
    normalized["productCode"] = normalized["productCode"].astype(str).str.strip()
    normalized["name"] = normalized["name"].fillna("").astype(str).str.strip()
    normalized["price"] = pd.to_numeric(normalized["price"], errors="coerce").fillna(0.0)

    normalized = normalized[normalized["productCode"] != ""]
    normalized = normalized.drop_duplicates(subset=["productCode"], keep="last")
    normalized = normalized.sort_values(["name", "productCode"]).reset_index(drop=True)
    return normalized


def compute_product_changes(current_df: pd.DataFrame, incoming_df: pd.DataFrame) -> pd.DataFrame:
    current = normalize_products_df(current_df)
    incoming = normalize_products_df(incoming_df)

    current_map = {
        str(row["productCode"]): row
        for _, row in current.iterrows()
    }
    incoming_map = {
        str(row["productCode"]): row
        for _, row in incoming.iterrows()
    }

    rows = []
    all_codes = sorted(set(current_map.keys()).union(incoming_map.keys()))
    for code in all_codes:
        old = current_map.get(code)
        new = incoming_map.get(code)

        old_price = float(old.get("price", 0.0)) if old is not None else 0.0
        new_price = float(new.get("price", 0.0)) if new is not None else 0.0
        old_name = str(old.get("name", "")) if old is not None else ""
        new_name = str(new.get("name", "")) if new is not None else ""
        name = new_name or old_name

        if old is None:
            change_type = "Novo"
        elif new is None:
            change_type = "Removido"
        elif abs(old_price - new_price) > 0.009:
            change_type = "Preço Alterado"
        elif old_name != new_name:
            change_type = "Nome Alterado"
        else:
            continue

        rows.append(
            {
                "productCode": code,
                "name": name,
                "old_price": old_price,
                "new_price": new_price,
                "diff_price": new_price - old_price,
                "change_type": change_type,
            }
        )

    return pd.DataFrame(rows)


def save_product_history(changes_df: pd.DataFrame, source_label: str) -> None:
    if changes_df.empty:
        return

    username = st.session_state.get("current_user", "unknown")
    if firestore_enabled():
        db = get_firestore_client()
        batch = db.batch()
        op_count = 0
        max_ops = 450

        for _, row in changes_df.iterrows():
            payload = {
                "productCode": str(row.get("productCode", "")),
                "name": str(row.get("name", "")),
                "old_price": float(row.get("old_price", 0.0)),
                "new_price": float(row.get("new_price", 0.0)),
                "diff_price": float(row.get("diff_price", 0.0)),
                "change_type": str(row.get("change_type", "")),
                "changed_by": username,
                "source": source_label,
                "changed_at": firestore.SERVER_TIMESTAMP,
            }
            ref = db.collection(FIRESTORE_PRODUCT_HISTORY_COLLECTION).document()
            batch.set(ref, payload)
            op_count += 1

            if op_count >= max_ops:
                batch.commit()
                batch = db.batch()
                op_count = 0

        if op_count:
            batch.commit()
        _fetch_product_history_firestore.clear()
    else:
        local_history = st.session_state.get("product_history_local", [])
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for _, row in changes_df.iterrows():
            local_history.append(
                {
                    "productCode": str(row.get("productCode", "")),
                    "name": str(row.get("name", "")),
                    "old_price": float(row.get("old_price", 0.0)),
                    "new_price": float(row.get("new_price", 0.0)),
                    "diff_price": float(row.get("diff_price", 0.0)),
                    "change_type": str(row.get("change_type", "")),
                    "changed_by": username,
                    "source": source_label,
                    "changed_at": now_str,
                }
            )
        st.session_state.product_history_local = local_history


@st.cache_data(show_spinner=False, ttl=300)
def _fetch_product_history_firestore(limit: int) -> pd.DataFrame:
    db = get_firestore_client()
    if not db:
        return pd.DataFrame()
    rows = []
    try:
        docs = (
            db.collection(FIRESTORE_PRODUCT_HISTORY_COLLECTION)
            .order_by("changed_at", direction=firestore.Query.DESCENDING)
            .limit(limit)
            .stream()
        )
    except Exception:
        docs = db.collection(FIRESTORE_PRODUCT_HISTORY_COLLECTION).stream()
    for doc in docs:
        data = doc.to_dict() or {}
        data["__doc_id"] = doc.id
        rows.append(data)
    if not rows:
        return pd.DataFrame()
    history_df = pd.DataFrame(rows)
    if "changed_at" in history_df.columns:
        history_df["changed_at_dt"] = history_df["changed_at"].apply(coerce_datetime)
        history_df = history_df.sort_values("changed_at_dt", ascending=False, na_position="last")
    return history_df


def fetch_product_history(limit: int = 300) -> pd.DataFrame:
    if firestore_enabled():
        return _fetch_product_history_firestore(limit)

    local_history = st.session_state.get("product_history_local", [])
    if not local_history:
        return pd.DataFrame()

    history_df = pd.DataFrame(local_history)
    if "changed_at" in history_df.columns:
        history_df["changed_at_dt"] = pd.to_datetime(history_df["changed_at"], errors="coerce")
        history_df = history_df.sort_values("changed_at_dt", ascending=False, na_position="last")
    return history_df


def save_setores_history(old_df: pd.DataFrame, new_df: pd.DataFrame) -> None:
    username = st.session_state.get("current_user", "unknown")
    old_map = {str(r["CódUnidade"]).strip(): str(r.get("items__description", "")) for _, r in old_df.iterrows()}
    new_map = {str(r["CódUnidade"]).strip(): str(r.get("items__description", "")) for _, r in new_df.iterrows()}
    all_codes = set(old_map) | set(new_map)
    changes = []
    for code in all_codes:
        old_desc = old_map.get(code)
        new_desc = new_map.get(code)
        if old_desc is None:
            change_type = "Novo"
        elif new_desc is None:
            change_type = "Removido"
        elif old_desc != new_desc:
            change_type = "Descrição Alterada"
        else:
            continue
        changes.append({"codigo": code, "old_desc": old_desc or "", "new_desc": new_desc or "", "change_type": change_type})

    if not changes:
        return

    if firestore_enabled():
        db = get_firestore_client()
        batch = db.batch()
        op_count = 0
        for ch in changes:
            payload = {**ch, "changed_by": username, "changed_at": firestore.SERVER_TIMESTAMP}
            batch.set(db.collection(FIRESTORE_SETORES_HISTORY_COLLECTION).document(), payload)
            op_count += 1
            if op_count >= 450:
                batch.commit()
                batch = db.batch()
                op_count = 0
        if op_count:
            batch.commit()
        _fetch_setores_history_firestore.clear()


@st.cache_data(show_spinner=False, ttl=300)
def _fetch_setores_history_firestore(limit: int) -> pd.DataFrame:
    db = get_firestore_client()
    if not db:
        return pd.DataFrame()
    rows = []
    try:
        docs = (
            db.collection(FIRESTORE_SETORES_HISTORY_COLLECTION)
            .order_by("changed_at", direction=firestore.Query.DESCENDING)
            .limit(limit)
            .stream()
        )
    except Exception:
        docs = db.collection(FIRESTORE_SETORES_HISTORY_COLLECTION).stream()
    for doc in docs:
        data = doc.to_dict() or {}
        rows.append(data)
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    if "changed_at" in df.columns:
        df["changed_at_dt"] = df["changed_at"].apply(coerce_datetime)
        df = df.sort_values("changed_at_dt", ascending=False, na_position="last")
    return df


def fetch_setores_history(limit: int = 200) -> pd.DataFrame:
    if firestore_enabled():
        return _fetch_setores_history_firestore(limit)
    return pd.DataFrame()


def sync_firestore_reference_data() -> None:
    if not firestore_enabled():
        return

    if st.session_state.excel_data is None:
        st.session_state.excel_data = {}

    st.session_state.excel_data["Produtos"] = fetch_products_df()
    st.session_state.excel_data["Setor"] = fetch_setores_df()


def get_storage_path() -> Optional[str]:
    path = os.getenv("ORDERS_WORKBOOK_PATH")
    if path and path.strip():
        return path
    return None


def render_sidebar() -> None:
    storage_path = get_storage_path()

    with st.sidebar:
        st.header("Gerenciar dados")
        
        if firestore_enabled():
            st.success("✅ Dados do Firestore carregados automaticamente")
            st.divider()
            
            if st.button("🔄 Recarregar dados do Firestore", use_container_width=True):
                fetch_products_df.clear()
                fetch_setores_df.clear()
                sync_firestore_to_session()
                st.toast("Dados recarregados do Firestore.", icon="✅")

            if st.session_state.excel_data is not None:
                produtos_df = st.session_state.excel_data.get("Produtos", pd.DataFrame())
                setores_df = st.session_state.excel_data.get("Setor", pd.DataFrame())
                
                st.divider()
                st.caption("📤 Para atualizar Produtos/Setor, carregue Excel e envie:")
                
                uploaded = st.file_uploader("Carregar planilha (.xlsx)", type=["xlsx", "xls"])
                if uploaded is not None:
                    set_excel_data(uploaded.getvalue(), f"upload ({uploaded.name})")
                    st.rerun()
                
                if st.button("📤 Enviar Produtos/Setor para Firestore", use_container_width=True):
                    if produtos_df.empty or setores_df.empty:
                        st.error("Carregue uma planilha com Produtos e Setor antes de enviar.")
                    else:
                        upsert_collection_from_df(
                            FIRESTORE_PRODUCTS_COLLECTION,
                            produtos_df,
                            "productCode",
                        )
                        upsert_collection_from_df(
                            FIRESTORE_SETORES_COLLECTION,
                            setores_df,
                            "CódUnidade",
                        )
                        fetch_products_df.clear()
                        fetch_setores_df.clear()
                        sync_firestore_reference_data()
                        st.toast("Produtos e setores enviados ao Firestore.", icon="✅")
        else:
            st.warning("⚠️ Firestore não disponível. Carregue uma planilha:")
            uploaded = st.file_uploader("Carregar planilha (.xlsx)", type=["xlsx", "xls"])
            if uploaded is not None:
                set_excel_data(uploaded.getvalue(), f"upload ({uploaded.name})")

        if storage_path:
            st.divider()
            st.caption(f"Arquivo do servidor: {storage_path}")
            if st.button("Carregar do servidor", use_container_width=True):
                if os.path.exists(storage_path):
                    with open(storage_path, "rb") as handler:
                        set_excel_data(handler.read(), os.path.basename(storage_path))
                else:
                    st.error("Arquivo configurado não encontrado no servidor.")


def build_workbook_bytes(data: Dict[str, pd.DataFrame]) -> io.BytesIO:
    """Build workbook using 'Excel Form Base.xlsx' as template if available."""
    from openpyxl import load_workbook
    
    # Try to load template
    model_path = os.path.join(os.path.dirname(__file__), "Excel Form Base.xlsx")
    if os.path.exists(model_path):
        # Load template and update sheets
        wb = load_workbook(model_path)
        
        # Clear existing data (keep headers)
        for sheet_name in ["Produtos", "Setor", "Aguardando Aprovação", "Pedido"]:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Delete all rows except the first (header)
                while ws.max_row > 1:
                    ws.delete_rows(2)
        
        # Write new data
        for sheet_name in ["Produtos", "Setor", "Aguardando Aprovação", "Pedido"]:
            if sheet_name in wb.sheetnames:
                df = data.get(sheet_name, pd.DataFrame())
                df = df.drop(columns=["__doc_id"], errors="ignore")
                
                ws = wb[sheet_name]
                # Write data starting from row 2 (after header)
                for r_idx, row in enumerate(df.itertuples(index=False), start=2):
                    for c_idx, value in enumerate(row, start=1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    else:
        # Fallback: create from scratch
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            for sheet_name in ["Produtos", "Setor", "Aguardando Aprovação", "Pedido"]:
                df = data.get(sheet_name, pd.DataFrame())
                df = df.drop(columns=["__doc_id"], errors="ignore")
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        buffer.seek(0)
        return buffer


def format_currency(value: float) -> str:
    return f"R$ {value:,.2f}".replace(",", "tmp").replace(".", ",").replace("tmp", ".")


def apply_layout_theme(use_new_layout: bool) -> None:
    if not use_new_layout:
        return

    st.markdown(
        """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

    :root{
      --bg:#F5F7FA;
      --surface:#FFFFFF;
      --title:#343C6A;
      --text:#4F4F4F;
      --muted:#718EBF;
      --border:#E6EFF5;
      --primary:#2D60FF;
    }

    /* APP */
    .stApp{
      background: var(--bg);
      font-family: Inter, system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif;
      color: var(--text);
    }

        .stApp, .stApp *{
            font-family: Inter, system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif !important;
        }

        /* 1) Controla o "respiro" geral do conteúdo */
        [data-testid="stAppViewContainer"] > .main .block-container{
            max-width: 1200px;
            padding-top: 24px;
            padding-bottom: 32px;
        }

        /* 2) O maior vilão: o spacing interno automático do Streamlit */
        [data-testid="stVerticalBlock"]{
            gap: 14px !important;      /* “grid gap” padrão entre elementos */
        }
        [data-testid="stVerticalBlock"] > div{
            margin: 0 !important;      /* remove margin “fantasma” */
        }

        /* 3) Títulos do Streamlit têm margem grande. Ajusta isso */
        h1{ font-size: 28px !important; margin: 0 0 14px 0 !important; font-weight: 700 !important; }
        h2{ font-size: 20px !important; margin: 0 0 10px 0 !important; font-weight: 700 !important; }
        h3{ font-size: 18px !important; margin: 0 0 8px 0 !important;  font-weight: 700 !important; }

        p, label, .stMarkdown{ margin: 0 !important; }

        /* 4) Labels do Streamlit (Selectbox/Number/Text) criam “altura extra” */
        [data-testid="stWidgetLabel"]{
            margin-bottom: 4px !important;
            color: #718EBF !important;
            font-weight: 600 !important;
            font-size: 13px !important;
        }

        /* 5) Inputs com altura e radius padrão (BankDash) */
        [data-testid="stTextInput"] input,
        [data-testid="stNumberInput"] input{
            height: 46px !important;
            border-radius: 12px !important;
        }
        [data-testid="stSelectbox"] > div{
            min-height: 46px !important;
            border-radius: 12px !important;
        }

    /* Remove travas / centraliza com max-width (NÃO use min-width fixa) */
    [data-testid="stAppViewContainer"] > .main .block-container{
      max-width: 1200px;
      padding-top: 24px;
      padding-bottom: 32px;
    }

    /* SIDEBAR - BankDash é clara */
    section[data-testid="stSidebar"]{
      background: var(--surface) !important;
      border-right: 1px solid var(--border);
    }
    section[data-testid="stSidebar"] *{
      color: var(--title);
    }

    /* TOPBAR */
    .dashboard-topbar{
      height: 88px;
      display:flex;
      align-items:center;
      gap:16px;
      background: var(--surface);
      border-bottom: 1px solid var(--border);
      padding: 0 24px;
      margin: -24px -24px 24px -24px; /* encosta na borda do container */
      border-radius: 18px;
    }

    .dashboard-title{
      margin:0;
      font-size:20px;
      font-weight:600;
      color:var(--title);
    }

    /* Search input estilo BankDash */
    [data-testid="stTextInput"] input{
      background: var(--bg) !important;
      border: 1px solid transparent !important;
      border-radius: 999px !important;
      height: 46px !important;
      padding-left: 16px !important;
    }
    [data-testid="stTextInput"] input:focus{
      border-color: #D9E3EF !important;
      box-shadow: 0 0 0 3px rgba(45,96,255,.12) !important;
    }

    /* Buttons redondinhos */
    .dashboard-topbar .stButton > button{
      width:46px !important;
      height:46px !important;
      border-radius:999px !important;
      background: var(--bg) !important;
      border: 1px solid var(--border) !important;
      box-shadow:none !important;
      padding:0 !important;
    }

    /* Cards padrão */
    .bd-card{
      background: var(--surface);
      border: 1px solid var(--border);
      border-radius: 20px;
      padding: 18px;
    }

    /* Títulos */
    h1,h2,h3{
      color: var(--title) !important;
      letter-spacing: -0.01em;
    }
    .bd-page-title{
      font-size: 28px;
      font-weight: 700;
      color: var(--title);
      margin: 8px 0 18px 0;
    }
    .bd-section-title{
      font-size: 20px;
      font-weight: 700;
      color: var(--title);
      margin: 0 0 10px 0;
    }

    /* Dataframe arredondado */
    [data-testid="stDataFrame"]{
      border: 1px solid var(--border);
      border-radius: 20px;
      overflow: hidden;
      background: var(--surface);
    }

        /* ====== BANKDASH CARD LAYOUT ====== */
        .bd-card{
            background: var(--surface);
            border: 1px solid var(--border);
            border-radius: 20px;
            padding: 22px 22px;
        }

        /* Remove espaços gigantes dentro do card (Streamlit coloca margem grande) */
        .bd-card .stMarkdown, .bd-card .stText, .bd-card .stCaption, .bd-card .stAlert{
            margin-top: 0 !important;
        }
        .bd-card [data-testid="stVerticalBlock"]{
            gap: 10px !important; /* controla espaçamento interno */
        }

        /* Título do card tipo BankDash */
        .bd-card-title{
            display:flex;
            align-items:center;
            justify-content:space-between;
            margin-bottom: 10px;
        }
        .bd-card-title h3{
            font-size: 18px;
            font-weight: 700;
            color: var(--title);
            margin: 0;
        }

        /* Inputs padrão mais "clean" */
        [data-testid="stTextInput"] input,
        [data-testid="stNumberInput"] input{
            background: #F5F7FA !important;
            border: 1px solid #E6EFF5 !important;
            border-radius: 12px !important;
            height: 46px !important;
        }
        [data-testid="stTextInput"] input:focus,
        [data-testid="stNumberInput"] input:focus{
            border-color: #D9E3EF !important;
            box-shadow: 0 0 0 3px rgba(45,96,255,.10) !important;
        }

        /* Selectbox */
        [data-testid="stSelectbox"] > div{
            background: #F5F7FA !important;
            border: 1px solid #E6EFF5 !important;
            border-radius: 12px !important;
            min-height: 46px !important;
        }

        /* Botão primário BankDash */
        .bd-primary .stButton > button{
            background: var(--primary) !important;
            border: 1px solid var(--primary) !important;
            color: white !important;
            border-radius: 12px !important;
            height: 46px !important;
            font-weight: 700 !important;
        }

        /* Botão secundário */
        .bd-secondary .stButton > button{
            background: #F5F7FA !important;
            border: 1px solid #E6EFF5 !important;
            color: var(--title) !important;
            border-radius: 12px !important;
            height: 46px !important;
            font-weight: 600 !important;
        }

        /* Mini tabela do carrinho (colunas) */
        .bd-table-head{
            font-size: 12px;
            color: var(--muted);
            font-weight: 700;
            text-transform: uppercase;
        }
        .bd-divider{
            height: 1px;
            background: var(--border);
            margin: 10px 0 12px 0;
        }

    </style>
    """,
        unsafe_allow_html=True,
    )


@st.cache_resource
def get_company_logo_path() -> Optional[str]:
    base_dir = os.path.dirname(__file__)
    for relative_path in LOGO_CANDIDATE_FILES:
        logo_path = os.path.join(base_dir, relative_path)
        if os.path.exists(logo_path):
            return logo_path
    return None


def render_company_branding() -> None:
    logo_path = get_company_logo_path()
    if logo_path:
        st.image(logo_path, use_container_width=True)
    else:
        st.markdown(f"### 🏢 {COMPANY_NAME}")


@st.cache_resource
def get_user_avatar_path() -> Optional[str]:
    base_dir = os.path.dirname(__file__)
    for relative_path in AVATAR_CANDIDATE_FILES:
        avatar_path = os.path.join(base_dir, relative_path)
        if os.path.exists(avatar_path):
            return avatar_path
    return None


def render_new_layout_header() -> None:
    st.markdown('<div class="dashboard-topbar">', unsafe_allow_html=True)
    st.markdown('<p class="dashboard-overview-label">Overview</p>', unsafe_allow_html=True)

    col_title, col_search, col_cfg, col_alert, col_avatar = st.columns([5.4, 2.55, 0.8, 0.8, 0.9])

    with col_title:
        st.markdown('<p class="dashboard-title">Overview</p>', unsafe_allow_html=True)

    with col_search:
        st.text_input(
            "Buscar geral",
            key="dashboard_global_search",
            placeholder="Search for something",
            label_visibility="collapsed",
        )

    with col_cfg:
        if st.button("⚙️", key="dashboard_settings_btn", help="Configurações"):
            st.toast("Configurações em breve", icon="⚙️")

    with col_alert:
        if st.button("🔔", key="dashboard_alert_btn", help="Alertas"):
            st.toast("Sem novos alertas", icon="🔔")

    with col_avatar:
        avatar_path = get_user_avatar_path()
        if avatar_path:
            st.image(avatar_path, width=50)
        else:
            username = str(st.session_state.get("current_user") or "U")
            st.markdown(f"### 👤 {username[:1].upper()}")

    st.markdown("</div>", unsafe_allow_html=True)


def parse_float(value: object) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def parse_int(value: object) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def coerce_datetime(value: object) -> Optional[datetime]:
    if value is None:
        return None
    if hasattr(value, "to_datetime"):
        try:
            return value.to_datetime()
        except Exception:
            return None
    if isinstance(value, datetime):
        return value
    try:
        parsed = pd.to_datetime(value, errors="coerce")
        if pd.isna(parsed):
            return None
        return parsed.to_pydatetime()
    except Exception:
        return None


def normalize_order_items(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["codigo", "nome", "qtde", "preco", "total"])

    normalized = pd.DataFrame()
    normalized["codigo"] = df.get("CódProImpakto", pd.Series(dtype=str)).astype(str)
    normalized["nome"] = df.get("Item", pd.Series(dtype=str)).fillna("").astype(str)
    normalized["qtde"] = pd.to_numeric(df.get("Qtde", 0), errors="coerce").fillna(0).astype(int)
    normalized["preco"] = pd.to_numeric(df.get("$ Unitário", 0.0), errors="coerce").fillna(0.0)
    normalized["total"] = pd.to_numeric(df.get("$ Total", 0.0), errors="coerce").fillna(0.0)

    # Preserve input sequence to keep PDF comparison aligned with the new order.
    grouped = (
        normalized.groupby(["codigo", "nome", "preco"], as_index=False, sort=False)
        .agg({"qtde": "sum", "total": "sum"})
    )
    return grouped


def fetch_latest_history_order_items_for_client(setor_desc: str) -> pd.DataFrame:
    db = get_firestore_client()
    if not db:
        return pd.DataFrame()

    try:
        target_desc = str(setor_desc).strip().lower()
        history_rows = []

        # Try server-side filter first (fast path — avoids full collection scan)
        for field in ("client_name", "SETOR2"):
            try:
                filtered_docs = (
                    db.collection("historico_pedidos")
                    .where(field, "==", str(setor_desc).strip())
                    .stream()
                )
                for doc in filtered_docs:
                    data = doc.to_dict() or {}
                    client_name = str(data.get("client_name") or data.get("SETOR2") or "").strip().lower()
                    if client_name == target_desc:
                        approved_dt = coerce_datetime(data.get("approved_at"))
                        export_dt = coerce_datetime(data.get("export_date"))
                        created_dt = coerce_datetime(data.get("created_at"))
                        sort_dt = approved_dt or export_dt or created_dt
                        history_rows.append((sort_dt, data))
            except Exception:
                pass
            if history_rows:
                break

        # Fallback: full scan (covers legacy records with different casing/fields)
        if not history_rows:
            for doc in db.collection("historico_pedidos").stream():
                data = doc.to_dict() or {}
                client_name = str(data.get("client_name") or data.get("SETOR2") or "").strip().lower()
                if client_name != target_desc:
                    continue
                approved_dt = coerce_datetime(data.get("approved_at"))
                export_dt = coerce_datetime(data.get("export_date"))
                created_dt = coerce_datetime(data.get("created_at"))
                sort_dt = approved_dt or export_dt or created_dt
                history_rows.append((sort_dt, data))

        if not history_rows:
            return pd.DataFrame()

        history_rows.sort(key=lambda item: item[0] or datetime.min, reverse=True)
        latest_history = history_rows[0][1]

        items = latest_history.get("items", []) or []
        if not items:
            return pd.DataFrame()

        rows = []
        for item in items:
            rows.append(
                {
                    "CódProImpakto": item.get("CódProImpakto"),
                    "Item": item.get("Item"),
                    "Qtde": item.get("Qtde"),
                    "$ Unitário": item.get("$ Unitário"),
                    "$ Total": item.get("$ Total"),
                    "approved_at": latest_history.get("approved_at"),
                    "order_number": latest_history.get("order_number"),
                    "SETOR2": latest_history.get("client_name"),
                    "Setor": latest_history.get("Setor"),
                }
            )

        return pd.DataFrame(rows)
    except Exception:
        return pd.DataFrame()


def fetch_latest_approved_order_items_for_setor(setor_codigo: str, setor_desc: str) -> pd.DataFrame:
    # Prefer historical orders as the comparison source.
    history_reference = fetch_latest_history_order_items_for_client(setor_desc)
    if not history_reference.empty:
        return history_reference

    if firestore_enabled():
        approved = fetch_firestore_rows_raw("pedido")
    else:
        approved = st.session_state.excel_data.get("Pedido", pd.DataFrame())

    if approved.empty:
        return pd.DataFrame()

    # Match by client description only to avoid coupling historical comparison to client IDs.
    desc_mask = approved.get("SETOR2", pd.Series(dtype=str)).astype(str).str.strip().str.lower() == str(setor_desc).strip().lower()
    approved = approved[desc_mask].copy()
    if approved.empty:
        return pd.DataFrame()

    approved["__approved_dt"] = approved.get("approved_at").apply(coerce_datetime)
    approved = approved.sort_values("__approved_dt", ascending=False, na_position="last")

    first_order = approved.iloc[0].get("order_number")
    if pd.notna(first_order):
        return approved[approved.get("order_number") == first_order].copy()

    return approved.head(1).copy()


def fetch_previous_approved_order_for_setor(setor_codigo: str, setor_desc: str) -> pd.DataFrame:
    return fetch_latest_approved_order_items_for_setor(setor_codigo, setor_desc)


def cart_to_order_dataframe(cart: List[Dict]) -> pd.DataFrame:
    if not cart:
        return pd.DataFrame()

    rows = []
    for item in cart:
        rows.append(
            {
                "CódProImpakto": item.get("codigo"),
                "Item": item.get("nome"),
                "Qtde": item.get("quantidade"),
                "$ Unitário": item.get("preco_unitario"),
                "$ Total": item.get("total"),
            }
        )
    return pd.DataFrame(rows)


def build_order_comparison(current_items: pd.DataFrame, previous_items: pd.DataFrame) -> pd.DataFrame:
    current_norm = normalize_order_items(current_items)
    prev_norm = normalize_order_items(previous_items)

    current_map = {str(row["codigo"]): row for _, row in current_norm.iterrows()}
    prev_map = {str(row["codigo"]): row for _, row in prev_norm.iterrows()}

    rows = []
    current_codes = [str(code) for code in current_norm.get("codigo", pd.Series(dtype=str)).tolist()]
    previous_codes = [str(code) for code in prev_norm.get("codigo", pd.Series(dtype=str)).tolist()]

    # Show current order items first, then historical-only items, without flipping sequence.
    all_codes = list(dict.fromkeys(current_codes + previous_codes))
    for code in all_codes:
        cur = current_map.get(code)
        prev = prev_map.get(code)

        nome = ""
        if cur is not None and str(cur.get("nome", "")).strip():
            nome = str(cur.get("nome", ""))
        elif prev is not None:
            nome = str(prev.get("nome", ""))

        qtde_atual = int(cur.get("qtde", 0)) if cur is not None else 0
        qtde_anterior = int(prev.get("qtde", 0)) if prev is not None else 0
        total_atual = float(cur.get("total", 0.0)) if cur is not None else 0.0
        total_anterior = float(prev.get("total", 0.0)) if prev is not None else 0.0

        if prev is None:
            status = "Novo"
        elif cur is None:
            status = "Removido"
        elif qtde_atual != qtde_anterior or abs(total_atual - total_anterior) > 0.01:
            status = "Alterado"
        else:
            status = "Igual"

        rows.append(
            {
                "Código": code,
                "Produto": nome,
                "Qtde Mês Anterior": qtde_anterior,
                "Qtde Atual": qtde_atual,
                "Δ Qtde": qtde_atual - qtde_anterior,
                "Total Mês Anterior": total_anterior,
                "Total Atual": total_atual,
                "Δ Total": total_atual - total_anterior,
                "Status": status,
            }
        )

    return pd.DataFrame(rows)


def load_latest_approved_order_into_cart() -> None:
    setor = st.session_state.selected_setor
    if not setor:
        st.warning("Selecione um setor antes de repetir o pedido anterior.")
        return

    previous = fetch_latest_approved_order_items_for_setor(str(setor.get("codigo")), str(setor.get("descricao")))
    if previous.empty:
        st.info("Nenhum pedido aprovado anterior encontrado para este setor.")
        return

    normalized = normalize_order_items(previous)
    if normalized.empty:
        st.info("O pedido anterior não possui itens válidos para repetição.")
        return

    st.session_state.cart = [
        {
            "codigo": str(row["codigo"]),
            "nome": str(row["nome"]),
            "quantidade": int(row["qtde"]),
            "preco_unitario": float(row["preco"]),
            "total": float(row["total"]),
        }
        for _, row in normalized.iterrows()
    ]

    st.success(f"✅ Pedido anterior carregado no carrinho com {len(st.session_state.cart)} item(ns).")


def render_pending_vs_previous_comparison(group_df: pd.DataFrame, group_id: str) -> None:
    if group_df.empty:
        return

    setor_codigo = str(group_df.iloc[0].get("Setor") or group_df.iloc[0].get("Unidade") or "")
    setor_desc = str(group_df.iloc[0].get("SETOR2") or "")

    previous = fetch_previous_approved_order_for_setor(setor_codigo, setor_desc)
    with st.expander("📊 Comparação com último pedido aprovado", expanded=False):
        if previous.empty:
            st.info("Não há pedido aprovado anterior para este setor.")
            return

        comparison = build_order_comparison(group_df, previous)
        if comparison.empty:
            st.info("Não foi possível gerar a comparação para este pedido.")
            return

        previous_date = coerce_datetime(previous.iloc[0].get("approved_at"))
        previous_order_number = previous.iloc[0].get("order_number")
        if previous_date:
            st.caption(
                f"Referência: pedido #{previous_order_number} aprovado em {previous_date.strftime('%d/%m/%Y %H:%M')}"
            )

        total_prev = float(comparison["Total Mês Anterior"].sum())
        total_current = float(comparison["Total Atual"].sum())
        delta_total = total_current - total_prev

        col_a, col_b, col_c = st.columns(3)
        col_a.metric("Total pedido anterior", format_currency(total_prev))
        col_b.metric("Total atual", format_currency(total_current))
        col_c.metric("Diferença", format_currency(delta_total))

        st.dataframe(comparison, use_container_width=True, hide_index=True)


def render_setor_selector(setor_df: pd.DataFrame) -> None:
    if setor_df.empty:
        st.error("A aba 'Setor' está vazia na planilha.")
        return

    required_cols = {"CódUnidade", "items__description"}
    missing = required_cols.difference(setor_df.columns)
    if missing:
        st.error(
            "A aba 'Setor' precisa conter as colunas: "
            + ", ".join(sorted(required_cols))
        )
        return

    setores = setor_df.dropna(subset=["CódUnidade", "items__description"]).copy()
    setores["codigo"] = setores["CódUnidade"].apply(parse_int)
    setores["descricao"] = setores["items__description"].astype(str)
    setores["label"] = setores.apply(
        lambda row: f"{row['codigo']} - {row['descricao']}", axis=1
    )

    if setores.empty:
        st.warning("Nenhum setor disponível.")
        return

    current_label = st.session_state.selected_setor.get("label") if st.session_state.selected_setor else None
    index = 0
    if current_label:
        try:
            index = setores["label"].tolist().index(current_label)
        except ValueError:
            index = 0

    if current_label and "search_setor" not in st.session_state:
        st.session_state["search_setor"] = current_label

    def setor_search(term: str) -> List[str]:
        if not term:
            return setores["label"].tolist()[:50]
        term_lower = term.lower()
        subset = setores.copy()
        subset["desc_lower"] = subset["descricao"].astype(str).str.lower()
        subset["code_lower"] = subset["codigo"].astype(str).str.lower()
        mask = (
            subset["desc_lower"].str.contains(term_lower)
            | subset["code_lower"].str.contains(term_lower)
        )
        subset = subset.loc[mask].copy()
        if subset.empty:
            return []
        subset["rank"] = (
            subset["desc_lower"].str.startswith(term_lower).astype(int) * 2
            + subset["code_lower"].str.startswith(term_lower).astype(int)
        )
        subset = subset.sort_values(["rank", "descricao"], ascending=[False, True])
        return subset["label"].tolist()[:50]

    selected_label = st_searchbox(
        setor_search,
        key="search_setor",
        label="2. Selecione o setor/cliente (digite para buscar)",
    )
    if not selected_label:
        st.info("Digite para buscar e selecione um setor/cliente.")
        return

    selected_row = setores[setores["label"] == selected_label].iloc[0]
    st.session_state.selected_setor = {
        "codigo": selected_row["codigo"],
        "descricao": selected_row["descricao"],
        "label": selected_row["label"],
    }



def render_product_selector(produtos_df: pd.DataFrame) -> Optional[dict]:
    if produtos_df.empty:
        st.error("A aba 'Produtos' está vazia na planilha.")
        return None

    required_cols = {"productCode", "name", "price"}
    missing = required_cols.difference(produtos_df.columns)
    if missing:
        st.error(
            "A aba 'Produtos' precisa conter as colunas: "
            + ", ".join(sorted(required_cols))
        )
        return None

    produtos = produtos_df.copy()
    produtos["price"] = pd.to_numeric(produtos.get("price"), errors="coerce").fillna(0.0)

    produtos = produtos.sort_values("name")
    produtos["label"] = produtos.apply(
        lambda row: f"{row['productCode']} - {row['name']}", axis=1
    )

    def produto_search(term: str) -> List[str]:
        if not term:
            return produtos["label"].tolist()[:50]
        term_lower = term.lower()
        subset = produtos.copy()
        subset["name_lower"] = subset["name"].astype(str).str.lower()
        subset["code_lower"] = subset["productCode"].astype(str).str.lower()
        mask = (
            subset["name_lower"].str.contains(term_lower)
            | subset["code_lower"].str.contains(term_lower)
        )
        subset = subset.loc[mask].copy()
        if subset.empty:
            return []
        subset["rank"] = (
            subset["name_lower"].str.startswith(term_lower).astype(int) * 2
            + subset["code_lower"].str.startswith(term_lower).astype(int)
        )
        subset = subset.sort_values(["rank", "name"], ascending=[False, True])
        return subset["label"].tolist()[:50]

    selected_label = st_searchbox(
        produto_search,
        key="search_produto",
        label="Produto (digite para buscar)",
    )
    if not selected_label:
        st.info("Digite para buscar um produto.")
        return None

    selected_row = produtos[produtos["label"] == selected_label].iloc[0]
    return selected_row.to_dict()


def add_item_to_cart(product: dict, quantity: int) -> None:
    preco = parse_float(product.get("price"))
    codigo = str(product.get("productCode"))
    nome = str(product.get("name"))

    for item in st.session_state.cart:
        if item["codigo"] == codigo:
            item["quantidade"] += quantity
            item["total"] = item["quantidade"] * item["preco_unitario"]
            break
    else:
        st.session_state.cart.append(
            {
                "codigo": codigo,
                "nome": nome,
                "quantidade": quantity,
                "preco_unitario": preco,
                "total": preco * quantity,
            }
        )


def render_cart(compact: bool = True) -> None:
    cart = st.session_state.cart
    if not cart:
        st.info("Nenhum item no carrinho ainda.")
        return

    # Cabeçalho pequeno estilo dashboard
    if not compact:
        st.markdown('<div class="bd-section-title">Itens do pedido</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="bd-table-head">Itens do pedido</div>', unsafe_allow_html=True)
        st.markdown('<div class="bd-divider"></div>', unsafe_allow_html=True)

    header_cols = st.columns([1.2, 3.5, 1, 1, 1, 0.8])
    header_cols[0].markdown('<span class="bd-table-head">Código</span>', unsafe_allow_html=True)
    header_cols[1].markdown('<span class="bd-table-head">Produto</span>', unsafe_allow_html=True)
    header_cols[2].markdown('<span class="bd-table-head">Qtd</span>', unsafe_allow_html=True)
    header_cols[3].markdown('<span class="bd-table-head">Unit.</span>', unsafe_allow_html=True)
    header_cols[4].markdown('<span class="bd-table-head">Total</span>', unsafe_allow_html=True)
    header_cols[5].markdown('<span class="bd-table-head">Rem.</span>', unsafe_allow_html=True)

    for idx, item in enumerate(cart):
        col_codigo, col_nome, col_qtd, col_preco, col_total, col_remove = st.columns(
            [1.2, 3.5, 1, 1, 1, 0.8]
        )
        col_codigo.write(item["codigo"])
        col_nome.write(item["nome"])

        nova_qtd = col_qtd.number_input(
            "Quantidade",
            min_value=1,
            max_value=9999,
            value=int(item["quantidade"]),
            key=f"cart_qty_{idx}",
            label_visibility="collapsed",
        )
        if nova_qtd != item["quantidade"]:
            item["quantidade"] = int(nova_qtd)
            item["total"] = item["quantidade"] * item["preco_unitario"]

        col_preco.write(format_currency(item["preco_unitario"]))
        col_total.write(format_currency(item["total"]))

        if col_remove.button("🗑️", key=f"remove_{idx}"):
            st.session_state.cart.pop(idx)
            st.rerun()

    total = sum(entry["total"] for entry in cart)

    # Rodapé tipo dashboard (sem metric gigante)
    st.markdown('<div class="bd-divider"></div>', unsafe_allow_html=True)
    c1, c2 = st.columns([3, 1])
    c1.markdown('<span class="bd-table-head">Total</span>', unsafe_allow_html=True)
    c2.markdown(f"**{format_currency(total)}**")


def build_cart_rows(destino: str) -> pd.DataFrame:
    setor = st.session_state.selected_setor
    cart = st.session_state.cart
    if not setor:
        st.warning("Selecione um setor antes de salvar o pedido.")
        return pd.DataFrame()
    if not cart:
        st.warning("Adicione itens ao carrinho antes de salvar.")
        return pd.DataFrame()

    linhas = []
    for item in cart:
        linhas.append(
            {
                "CòdClienteImpakto": COD_CLIENTE_IMPAKTO,
                "CódProImpakto": item["codigo"],
                "Item": item["nome"],
                "Qtde": int(item["quantidade"]),
                "$ Unitário": float(item["preco_unitario"]),
                "$ Total": float(item["total"]),
                "Unidade": setor["codigo"],
                "Setor": setor["codigo"],
                "SETOR2": setor["descricao"],
            }
        )
    return pd.DataFrame(linhas)


def persist_cart(destino: str) -> None:
    linhas = build_cart_rows(destino)
    if linhas.empty:
        return

    # Generate PDF for storage
    pdf_buffer = generate_previa_pdf()
    
    if firestore_enabled():
        status = "aguardando" if destino == "Aguardando Aprovação" else "pedido"
        pending_order_id = generate_pending_order_id() if status == "aguardando" else None
        save_rows_to_firestore(linhas, status, pdf_buffer, pending_order_id)
        sync_firestore_to_session()
    else:
        destino_df = st.session_state.excel_data.get(destino, pd.DataFrame())
        destino_df = pd.concat([destino_df, linhas], ignore_index=True)
        st.session_state.excel_data[destino] = destino_df
    
    st.success(f"✅ Pedido do cliente '{st.session_state.selected_setor.get('descricao', 'cliente')}' salvo com sucesso!")
    st.session_state.cart = []
    st.session_state["_previa_pdf"] = None
    st.session_state["_previa_pdf_key"] = None
    st.success(
        f"{len(linhas)} item(ns) salvo(s) em '{destino}'. Atualize a aba correspondente para visualizar."
    )


def generate_previa_pdf() -> Optional[io.BytesIO]:
    cart = st.session_state.cart
    setor = st.session_state.selected_setor
    if not cart or not setor:
        return None

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle(
        "Titulo",
        parent=styles["Heading1"],
        fontSize=16,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#1a1a1a"),
    )
    elements.append(Paragraph("<b>PRÉVIA DE PEDIDO - IMPAKTO</b>", title_style))
    elements.append(Spacer(1, 12))

    info_table = Table(
        [
            ["Data:", datetime.now().strftime("%d/%m/%Y %H:%M")],
            ["Cliente:", str(COD_CLIENTE_IMPAKTO)],
            ["Setor:", f"{setor['codigo']} - {setor['descricao']}"],
            ["Itens:", str(len(cart))],
        ],
        colWidths=[3.5 * 28.35, 10 * 28.35],
    )
    info_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (0, -1), "RIGHT"),
                ("ALIGN", (1, 0), (1, -1), "LEFT"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    elements.append(info_table)
    elements.append(Spacer(1, 12))

    data = [["Código", "Produto", "Qtde", "Valor Unit.", "Total"]]
    total_pedido = 0.0
    for item in cart:
        data.append(
            [
                item["codigo"],
                item["nome"],
                str(item["quantidade"]),
                format_currency(item["preco_unitario"]),
                format_currency(item["total"]),
            ]
        )
        total_pedido += item["total"]

    data.append(["", "", "", "TOTAL", format_currency(total_pedido)])

    table = Table(data, colWidths=[2.5 * 28.35, 8 * 28.35, 1.5 * 28.35, 3 * 28.35, 3 * 28.35])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -2), 0.3, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F2F2F2")]),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
                ("LINEABOVE", (0, -1), (-1, -1), 1.5, colors.black),
            ]
        )
    )
    elements.append(table)
    elements.append(Spacer(1, 10))

    # Comparison with last approved order for the same setor
    try:
        previous_order = fetch_previous_approved_order_for_setor(
            str(setor.get("codigo")),
            str(setor.get("descricao")),
        )
        current_df = cart_to_order_dataframe(cart)
        comparison_df = build_order_comparison(current_df, previous_order)

        if not previous_order.empty and not comparison_df.empty:
            previous_date = coerce_datetime(previous_order.iloc[0].get("approved_at"))
            previous_order_number = previous_order.iloc[0].get("order_number")
            has_order_number = pd.notna(previous_order_number) and str(previous_order_number).strip().lower() not in {"", "none", "nan"}
            previous_label = f"Pedido #{previous_order_number}" if has_order_number else "Pedido sem numero"
            if previous_date:
                previous_label += f" ({previous_date.strftime('%d/%m/%Y %H:%M')})"

            subtitle_style = ParagraphStyle(
                "Subtitulo",
                parent=styles["Heading3"],
                fontSize=11,
                alignment=TA_LEFT,
                textColor=colors.HexColor("#1f2937"),
            )
            elements.append(Paragraph("<b>Comparacao com ultimo pedido aprovado</b>", subtitle_style))
            elements.append(Spacer(1, 6))
            elements.append(Paragraph(f"Referencia: {previous_label}", styles["Normal"]))
            elements.append(Spacer(1, 6))

            total_prev = float(comparison_df["Total Mês Anterior"].sum())
            total_current = float(comparison_df["Total Atual"].sum())
            delta_total = total_current - total_prev
            summary_table = Table(
                [
                    ["Total anterior", format_currency(total_prev)],
                    ["Total atual", format_currency(total_current)],
                    ["Diferenca", format_currency(delta_total)],
                ],
                colWidths=[4.5 * 28.35, 4.0 * 28.35],
            )
            summary_table.setStyle(
                TableStyle(
                    [
                        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 10),
                        ("ALIGN", (0, 0), (0, -1), "LEFT"),
                        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.HexColor("#F8FAFC"), colors.white]),
                        ("TOPPADDING", (0, 0), (-1, -1), 4),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ]
                )
            )
            elements.append(summary_table)
            elements.append(Spacer(1, 8))

            comparison_header_style = ParagraphStyle(
                "ComparacaoHeader",
                parent=styles["Normal"],
                fontName="Helvetica-Bold",
                fontSize=8,
                leading=9,
                alignment=TA_CENTER,
                textColor=colors.white,
            )
            comparison_cell_style = ParagraphStyle(
                "ComparacaoCell",
                parent=styles["Normal"],
                fontName="Helvetica",
                fontSize=8,
                leading=9,
                alignment=TA_LEFT,
                textColor=colors.HexColor("#1f2937"),
            )

            comparison_rows = [[
                Paragraph("Codigo", comparison_header_style),
                Paragraph("Produto", comparison_header_style),
                Paragraph("Qtde Ant.", comparison_header_style),
                Paragraph("Qtde Atual", comparison_header_style),
                Paragraph("Valor Ant.", comparison_header_style),
                Paragraph("Valor Atual", comparison_header_style),
                Paragraph("Dif. Valor", comparison_header_style),
                Paragraph("Status", comparison_header_style),
            ]]
            for _, row in comparison_df.iterrows():
                qtde_ant_val = pd.to_numeric(row.get("Qtde Mês Anterior", 0), errors="coerce")
                qtde_atual_val = pd.to_numeric(row.get("Qtde Atual", 0), errors="coerce")
                total_ant_val = pd.to_numeric(row.get("Total Mês Anterior", 0.0), errors="coerce")
                total_atual_val = pd.to_numeric(row.get("Total Atual", 0.0), errors="coerce")
                delta_val = pd.to_numeric(row.get("Δ Total", 0.0), errors="coerce")

                qtde_ant = int(0 if pd.isna(qtde_ant_val) else qtde_ant_val)
                qtde_atual = int(0 if pd.isna(qtde_atual_val) else qtde_atual_val)
                total_ant = float(0.0 if pd.isna(total_ant_val) else total_ant_val)
                total_atual_row = float(0.0 if pd.isna(total_atual_val) else total_atual_val)
                delta_row = float(0.0 if pd.isna(delta_val) else delta_val)

                comparison_rows.append(
                    [
                        str(row.get("Código", "")),
                        Paragraph(str(row.get("Produto", "")), comparison_cell_style),
                        str(qtde_ant),
                        str(qtde_atual),
                        format_currency(total_ant),
                        format_currency(total_atual_row),
                        format_currency(delta_row),
                        str(row.get("Status", "")),
                    ]
                )

            comparison_rows.append(
                [
                    "",
                    "TOTAL GERAL",
                    "",
                    "",
                    format_currency(total_prev),
                    format_currency(total_current),
                    format_currency(delta_total),
                    "",
                ]
            )

            comparison_table = Table(
                comparison_rows,
                colWidths=[1.35 * 28.35, 4.55 * 28.35, 1.2 * 28.35, 1.2 * 28.35, 1.55 * 28.35, 1.55 * 28.35, 1.55 * 28.35, 1.55 * 28.35],
            )
            comparison_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                        ("ALIGN", (0, 1), (0, -2), "CENTER"),
                        ("ALIGN", (1, 1), (1, -1), "LEFT"),
                        ("ALIGN", (2, 1), (3, -1), "CENTER"),
                        ("ALIGN", (4, 1), (6, -1), "RIGHT"),
                        ("ALIGN", (7, 1), (7, -1), "CENTER"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F8FAFC")]),
                        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E2E8F0")),
                        ("FONTNAME", (1, -1), (6, -1), "Helvetica-Bold"),
                        ("TOPPADDING", (0, 0), (-1, -1), 3),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ]
                )
            )
            elements.append(comparison_table)
            elements.append(Spacer(1, 10))
        else:
            elements.append(
                Paragraph(
                    "<i>Comparativo indisponivel para este cliente (sem pedido aprovado anterior compativel).</i>",
                    styles["Normal"],
                )
            )
            elements.append(Spacer(1, 8))
    except Exception as comparison_error:
        st.warning(f"⚠️ Comparativo não pôde ser incluído no PDF: {comparison_error}")
        elements.append(
            Paragraph(
                "<i>Comparativo indisponivel no momento devido a inconsistencias de dados.</i>",
                styles["Normal"],
            )
        )
        elements.append(Spacer(1, 8))

    note_style = ParagraphStyle(
        "Nota",
        parent=styles["Normal"],
        fontSize=9,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#666666"),
    )
    elements.append(
        Paragraph(
            "<i>Documento gerado automaticamente. Confirme o pedido antes de enviar.</i>",
            note_style,
        )
    )

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_history_order_pdf(order: Dict, client_label: str, setor_label: str) -> io.BytesIO:
    """Generate a printable PDF for a single order from history."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle(
        "TituloHistorico",
        parent=styles["Heading1"],
        fontSize=15,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#1a1a1a"),
    )
    elements.append(Paragraph("<b>HISTÓRICO DE PEDIDO - IMPAKTO</b>", title_style))
    elements.append(Spacer(1, 12))

    order_number = order.get("order_number", "N/A")
    created_at = order.get("created_at")
    approved_at = order.get("approved_at")
    export_date = order.get("export_date")

    def fmt_dt(value: object) -> str:
        if hasattr(value, "to_datetime"):
            value = value.to_datetime()
        if isinstance(value, datetime):
            return value.strftime("%d/%m/%Y %H:%M")
        return str(value) if value else "N/A"

    items = order.get("items", []) or []
    total = parse_float(order.get("$ Total"))
    if not total:
        total = sum(parse_float(item.get("$ Total")) for item in items)

    info_table = Table(
        [
            ["Pedido:", str(order_number)],
            ["Cliente:", str(client_label)],
            ["Setor:", str(setor_label)],
            ["Criado em:", fmt_dt(created_at)],
            ["Aprovado em:", fmt_dt(approved_at)],
            ["Exportado em:", fmt_dt(export_date)],
            ["Itens:", str(len(items))],
            ["Total:", format_currency(total)],
        ],
        colWidths=[4.2 * 28.35, 9.3 * 28.35],
    )
    info_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (0, -1), "RIGHT"),
                ("ALIGN", (1, 0), (1, -1), "LEFT"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    elements.append(info_table)
    elements.append(Spacer(1, 12))

    table_data = [["Código", "Produto", "Qtde", "Valor Unit.", "Total", "Un."]]
    for item in items:
        table_data.append(
            [
                str(item.get("CódProImpakto", "")),
                str(item.get("Item", "")),
                str(parse_int(item.get("Qtde"))),
                format_currency(parse_float(item.get("$ Unitário"))),
                format_currency(parse_float(item.get("$ Total"))),
                str(item.get("Unidade", "")),
            ]
        )

    if len(table_data) == 1:
        table_data.append(["-", "Nenhum item registrado", "", "", "", ""])

    table_data.append(["", "", "", "TOTAL", format_currency(total), ""])

    items_table = Table(
        table_data,
        colWidths=[2.0 * 28.35, 6.6 * 28.35, 1.5 * 28.35, 2.7 * 28.35, 2.7 * 28.35, 1.5 * 28.35],
    )
    items_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E6EA6")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -2), 0.3, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F7F7F7")]),
                ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("LINEABOVE", (0, -1), (-1, -1), 1.2, colors.black),
            ]
        )
    )
    elements.append(items_table)
    elements.append(Spacer(1, 10))

    note_style = ParagraphStyle(
        "NotaHistorico",
        parent=styles["Normal"],
        fontSize=9,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#666666"),
    )
    elements.append(Paragraph("<i>Documento para impressão do histórico de pedidos.</i>", note_style))

    doc.build(elements)
    buffer.seek(0)
    return buffer


def format_datetime_display(value: object) -> str:
    converted = coerce_datetime(value)
    if converted is None:
        return str(value) if value else "N/A"
    return converted.strftime("%Y-%m-%d %H:%M:%S")


def extract_date_iso(value: object) -> str:
    converted = coerce_datetime(value)
    if converted is None:
        return ""
    return converted.strftime("%Y-%m-%d")


def generate_history_batch_pdf(selected_orders: List[Dict]) -> io.BytesIO:
    """Generate one printable PDF containing all selected orders (one order per page)."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    for idx, order in enumerate(selected_orders):
        client_label = str(order.get("client_name") or "Sem nome")
        setor_label = str(order.get("Setor") or "N/A")
        order_number = str(order.get("order_number") or order.get("__doc_id") or "N/A")
        items = order.get("items", []) or []
        total = parse_float(order.get("$ Total"))
        if not total:
            total = sum(parse_float(item.get("$ Total")) for item in items)

        title_style = ParagraphStyle(
            "TituloHistoricoBatch",
            parent=styles["Heading1"],
            fontSize=14,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#1a1a1a"),
        )
        elements.append(Paragraph("<b>HISTÓRICO DE PEDIDO - IMPAKTO</b>", title_style))
        elements.append(Spacer(1, 10))

        info_table = Table(
            [
                ["Pedido:", order_number],
                ["Cliente:", client_label],
                ["Setor:", setor_label],
                ["Criado em:", format_datetime_display(order.get("created_at"))],
                ["Aprovado em:", format_datetime_display(order.get("approved_at"))],
                ["Exportado em:", format_datetime_display(order.get("export_date"))],
                ["Itens:", str(len(items))],
                ["Total:", format_currency(total)],
            ],
            colWidths=[4.2 * 28.35, 9.3 * 28.35],
        )
        info_table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("ALIGN", (0, 0), (0, -1), "RIGHT"),
                    ("ALIGN", (1, 0), (1, -1), "LEFT"),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        elements.append(info_table)
        elements.append(Spacer(1, 10))

        table_data = [["Código", "Produto", "Qtde", "Valor Unit.", "Total", "Un."]]
        for item in items:
            table_data.append(
                [
                    str(item.get("CódProImpakto", "")),
                    str(item.get("Item", "")),
                    str(parse_int(item.get("Qtde"))),
                    format_currency(parse_float(item.get("$ Unitário"))),
                    format_currency(parse_float(item.get("$ Total"))),
                    str(item.get("Unidade", "")),
                ]
            )
        if len(table_data) == 1:
            table_data.append(["-", "Nenhum item registrado", "", "", "", ""])
        table_data.append(["", "", "", "TOTAL", format_currency(total), ""])

        items_table = Table(
            table_data,
            colWidths=[2.0 * 28.35, 6.6 * 28.35, 1.5 * 28.35, 2.7 * 28.35, 2.7 * 28.35, 1.5 * 28.35],
        )
        items_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E6EA6")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -2), 0.3, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F7F7F7")]),
                    ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ("LINEABOVE", (0, -1), (-1, -1), 1.2, colors.black),
                ]
            )
        )
        elements.append(items_table)

        if idx < len(selected_orders) - 1:
            elements.append(PageBreak())

    doc.build(elements)
    buffer.seek(0)
    return buffer


def _previa_cache_key() -> str:
    cart = st.session_state.get("cart", [])
    setor = st.session_state.get("selected_setor") or {}
    return str(hash((str(cart), str(setor.get("codigo", "")))))


def render_previa_download() -> None:
    cart = st.session_state.get("cart")
    setor = st.session_state.get("selected_setor")
    if not cart or not setor:
        return

    cache_key = _previa_cache_key()
    if st.session_state.get("_previa_pdf_key") != cache_key:
        st.session_state["_previa_pdf"] = generate_previa_pdf()
        st.session_state["_previa_pdf_key"] = cache_key

    pdf_buffer = st.session_state.get("_previa_pdf")
    if pdf_buffer is None:
        return

    st.info("📋 Prévia do pedido gerada")
    file_name = f"PREVIA_Pedido_{setor['codigo']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

    col1, col2 = st.columns(2)
    with col1:
        st.download_button(
            "📥 Baixar PDF (para imprimir)",
            data=pdf_buffer.getvalue(),
            file_name=file_name,
            mime="application/pdf",
            use_container_width=True,
        )
    with col2:
        st.caption("✅ Será também salvo no Storage ao clicar em 'Salvar'")


def render_save_buttons() -> None:
    col1, col2 = st.columns(2)
    disabled = not st.session_state.cart or not st.session_state.selected_setor
    if col1.button("Salvar em 'Aguardando Aprovação'", disabled=disabled):
        persist_cart("Aguardando Aprovação")
    if col2.button("Aprovar direto em 'Pedido'", disabled=disabled):
        persist_cart("Pedido")



def update_excel_snapshot() -> None:
    if st.session_state.excel_data is None:
        return
    buffer = build_workbook_bytes(st.session_state.excel_data)
    st.session_state.excel_bytes = buffer.getvalue()




def move_approved_to_history(order_docs: List) -> int:
    """Move approved orders (documents with status='pedido') to history.
    Groups by order_number (fallback to __doc_id for legacy records) and creates
    one history record per order.
    Returns the count of moved orders."""
    try:
        db = get_firestore_client()
        moved_count = 0
        
        # Group documents by order_number (legacy fallback: doc id)
        order_groups = {}
        for doc in order_docs:
            order_data = doc.to_dict()
            order_key = order_data.get("order_number") or doc.id
            if order_key not in order_groups:
                order_groups[order_key] = []
            order_groups[order_key].append((doc.id, order_data))
        
        # For each order group, create a history record
        for order_number, items_list in order_groups.items():
            export_date = datetime.now(timezone.utc)
            
            # Collect items data
            items_data = []
            for doc_id, order_data in items_list:
                items_data.append({
                    "CódProImpakto": order_data.get("CódProImpakto"),
                    "Item": order_data.get("Item"),
                    "Qtde": order_data.get("Qtde"),
                    "$ Unitário": order_data.get("$ Unitário"),
                    "$ Total": order_data.get("$ Total"),
                    "Unidade": order_data.get("Unidade"),
                })
            
            # Collect audit trail entries
            audit_entries = []
            for doc_id, order_data in items_list:
                try:
                    audit_docs = db.collection(FIRESTORE_AUDIT_COLLECTION)\
                        .where("doc_id", "==", doc_id)\
                        .where("action", "in", ["item_added", "item_removed", "quantity_changed"])\
                        .stream()
                    
                    for audit_doc in audit_docs:
                        audit_data = audit_doc.to_dict()
                        audit_entries.append({
                            "action": audit_data.get("action"),
                            "timestamp": audit_data.get("timestamp"),
                            "details": audit_data.get("details", {})
                        })
                except:
                    pass
            
            # Get total value
            total_value = sum(item.get("$ Total", 0) for item in items_data)
            
            # Get order-level info from first item
            first_item = items_list[0][1] if items_list else {}
            setor = first_item.get("Setor", "Unknown")
            client_name = first_item.get("SETOR2") or first_item.get("client_name")
            
            # Create history record
            history_record = {
                "order_number": order_number,
                "Setor": setor,
                "client_name": client_name,
                "$ Total": total_value,
                "created_at": first_item.get("created_at"),
                "approved_at": first_item.get("approved_at"),
                "approved_by": first_item.get("approved_by"),
                "export_date": export_date,
                "items": items_data,
                "audit_trail": audit_entries,
                "item_count": len(items_data)
            }
            
            # Save to history
            db.collection("historico_pedidos").add(history_record)
            
            # Delete all items for this order from pedido_itens
            for doc_id, _ in items_list:
                db.collection(FIRESTORE_COLLECTION).document(doc_id).delete()
            
            moved_count += 1
        
        return moved_count
        
    except Exception as e:
        st.error(f"❌ Erro ao mover pedidos para histórico: {e}")
        return 0


def repair_approved_orders_safe(order_docs: List, apply_changes: bool = False) -> Dict[str, object]:
    """Safely repair legacy approved items grouping.

    Rules:
    - Fix items with missing order_number only when pending_order_id exists.
    - Split mixed order_number groups only when multiple pending_order_id values exist.
    - Leave ambiguous items (without pending_order_id) unchanged.
    """
    result = {
        "total_approved_items": len(order_docs),
        "updates_planned": 0,
        "updated_items": 0,
        "missing_order_groups_fixed": 0,
        "mixed_order_groups_split": 0,
        "unresolved_items": 0,
        "applied": apply_changes,
    }

    if not order_docs:
        return result

    missing_by_pending: Dict[str, List[str]] = {}
    missing_without_pending: List[str] = []
    order_to_pending: Dict[str, Dict[str, List[str]]] = {}

    for doc in order_docs:
        data = doc.to_dict() or {}
        doc_id = doc.id
        order_number = data.get("order_number")
        pending_order_id = str(data.get("pending_order_id") or "").strip()

        if order_number is None or str(order_number).strip() == "":
            if pending_order_id:
                missing_by_pending.setdefault(pending_order_id, []).append(doc_id)
            else:
                missing_without_pending.append(doc_id)
            continue

        order_key = str(order_number).strip()
        if pending_order_id:
            order_to_pending.setdefault(order_key, {}).setdefault(pending_order_id, []).append(doc_id)

    updates: Dict[str, int] = {}

    for pending_order_id, doc_ids in missing_by_pending.items():
        new_order_number = get_next_order_number() if apply_changes else 1
        for doc_id in doc_ids:
            updates[doc_id] = new_order_number

    split_groups = 0
    for order_key, pending_map in order_to_pending.items():
        if len(pending_map) <= 1:
            continue

        keep_pending = max(pending_map.items(), key=lambda item: len(item[1]))[0]
        for pending_order_id, doc_ids in pending_map.items():
            if pending_order_id == keep_pending:
                continue
            new_order_number = get_next_order_number() if apply_changes else 1
            for doc_id in doc_ids:
                updates[doc_id] = new_order_number
            split_groups += 1

    result["updates_planned"] = len(updates)
    result["missing_order_groups_fixed"] = len(missing_by_pending)
    result["mixed_order_groups_split"] = split_groups
    result["unresolved_items"] = len(missing_without_pending)

    if not apply_changes or not updates:
        return result

    try:
        db = get_firestore_client()
        if not db:
            return result

        batch = db.batch()
        op_count = 0
        max_ops = 450

        for doc_id, new_order_number in updates.items():
            doc_ref = db.collection(FIRESTORE_COLLECTION).document(doc_id)
            batch.update(doc_ref, {
                "order_number": new_order_number,
                "updated_at": firestore.SERVER_TIMESTAMP,
            })
            op_count += 1

            if op_count >= max_ops:
                batch.commit()
                batch = db.batch()
                op_count = 0

        if op_count > 0:
            batch.commit()

        result["updated_items"] = len(updates)
        return result
    except Exception as exc:
        st.error(f"❌ Erro ao aplicar correção segura: {exc}")
        return result


def build_ambiguous_approved_items_df(order_docs: List) -> pd.DataFrame:
    rows = []
    for doc in order_docs:
        data = doc.to_dict() or {}
        pending_order_id = str(data.get("pending_order_id") or "").strip()
        if pending_order_id:
            continue

        rows.append(
            {
                "__doc_id": doc.id,
                "SETOR2": data.get("SETOR2"),
                "Setor": data.get("Setor"),
                "CódProImpakto": data.get("CódProImpakto"),
                "Item": data.get("Item"),
                "Qtde": data.get("Qtde"),
                "$ Unitário": data.get("$ Unitário"),
                "$ Total": data.get("$ Total"),
                "order_number": data.get("order_number"),
                "approved_at": data.get("approved_at"),
            }
        )

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    if "approved_at" in df.columns:
        df = df.sort_values(by="approved_at", ascending=False, na_position="last")
    return df


def render_download_section() -> None:
    if st.session_state.excel_data is None:
        return

    update_excel_snapshot()
    st.divider()
    st.subheader("Exportar planilha atualizada")
    excel_bytes = st.session_state.excel_bytes
    file_name = f"planilha_pedidos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    st.download_button(
        "Baixar Excel atualizado",
        data=excel_bytes,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    storage_path = get_storage_path()
    if storage_path:
        if st.button("Sobrescrever arquivo no servidor", use_container_width=True):
            dir_name = os.path.dirname(storage_path)
            if dir_name:
                os.makedirs(dir_name, exist_ok=True)
            with open(storage_path, "wb") as handler:
                handler.write(excel_bytes)
            st.success(f"Arquivo salvo em {storage_path}")
    
    # Move approved orders to history
    st.divider()
    st.subheader("Gerenciar Pedidos Aprovados")
    
    if firestore_enabled():
        try:
            db = get_firestore_client()
            approved_orders = list(db.collection(FIRESTORE_COLLECTION)\
                .where("status", "==", "pedido")\
                .stream())
            
            if approved_orders:
                st.info(f"📊 Existem {len(approved_orders)} item(ns) aprovado(s) aguardando movimentação para histórico.")

                analysis = repair_approved_orders_safe(approved_orders, apply_changes=False)
                if analysis["updates_planned"] > 0 or analysis["unresolved_items"] > 0:
                    st.caption(
                        "Correção segura (legado): "
                        f"{analysis['updates_planned']} item(ns) podem ser ajustados automaticamente; "
                        f"{analysis['unresolved_items']} item(ns) exigem revisão manual."
                    )

                ambiguous_df = build_ambiguous_approved_items_df(approved_orders)
                if not ambiguous_df.empty:
                    with st.expander(f"🔎 Revisão manual: {len(ambiguous_df)} item(ns) ambíguo(s)", expanded=False):
                        st.caption("Itens sem pending_order_id. Revise antes de mover para histórico.")
                        display_df = ambiguous_df.copy()
                        if "approved_at" in display_df.columns:
                            display_df["approved_at"] = pd.to_datetime(
                                display_df["approved_at"], errors="coerce"
                            ).dt.strftime("%Y-%m-%d %H:%M:%S")
                        st.dataframe(
                            display_df[[
                                "__doc_id",
                                "SETOR2",
                                "Setor",
                                "CódProImpakto",
                                "Item",
                                "Qtde",
                                "$ Unitário",
                                "$ Total",
                                "order_number",
                                "approved_at",
                            ]],
                            use_container_width=True,
                            hide_index=True,
                        )
                        st.download_button(
                            "📥 Baixar lista de itens ambíguos (CSV)",
                            data=display_df.to_csv(index=False).encode("utf-8-sig"),
                            file_name=f"itens_ambiguos_aprovados_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                            mime="text/csv",
                            use_container_width=True,
                        )

                col_fix, col_move = st.columns(2)
                with col_fix:
                    if st.button("🛠️ Corrigir aprovados legados (seguro)", use_container_width=True):
                        with st.spinner("Aplicando correção segura..."):
                            repair_result = repair_approved_orders_safe(approved_orders, apply_changes=True)

                        if repair_result["updated_items"] > 0:
                            st.success(
                                "✅ Correção aplicada: "
                                f"{repair_result['updated_items']} item(ns), "
                                f"{repair_result['mixed_order_groups_split']} grupo(s) dividido(s), "
                                f"{repair_result['missing_order_groups_fixed']} grupo(s) sem número corrigido(s)."
                            )
                            if repair_result["unresolved_items"] > 0:
                                st.warning(
                                    f"⚠️ {repair_result['unresolved_items']} item(ns) antigos continuam ambíguos e precisam revisão manual."
                                )
                            sync_firestore_to_session()
                            st.rerun()
                        else:
                            if repair_result["unresolved_items"] > 0:
                                st.warning(
                                    "Nenhuma correção automática aplicada. "
                                    f"{repair_result['unresolved_items']} item(ns) são ambíguos (sem pending_order_id)."
                                )
                            else:
                                st.info("Nenhum ajuste legado necessário.")
                
                with col_move:
                    if st.button("📚 Mover todos os pedidos aprovados para histórico", use_container_width=True):
                        with st.spinner("Movendo pedidos para histórico..."):
                            moved_count = move_approved_to_history(approved_orders)
                            
                            if moved_count > 0:
                                st.success(f"✅ {moved_count} pedido(s) movido(s) para histórico com sucesso!")
                                sync_firestore_to_session()
                                st.rerun()
            else:
                st.info("✅ Nenhum pedido aprovado aguardando movimentação.")
        except Exception as e:
            st.warning(f"⚠️ Erro ao verificar pedidos aprovados: {e}")


def render_aguardando_tab() -> None:
    current_user = st.session_state.get("current_user")
    user_info = get_user_info(current_user)
    can_edit_aguardando = user_info.get("role") in {"admin", "user", "catalog_editor"}

    if firestore_enabled():
        df = fetch_firestore_rows_raw("aguardando")
    else:
        df = st.session_state.excel_data.get("Aguardando Aprovação", pd.DataFrame())
    
    # Header with refresh button
    col_header1, col_header2 = st.columns([3, 1])
    with col_header1:
        st.subheader("📋 Pedidos aguardando aprovação")
    with col_header2:
        if st.button("🔄 Atualizar", key="refresh_aguardando", help="Atualizar dados do Firestore"):
            sync_firestore_to_session()
            st.success("✅ Dados atualizados!")
            st.rerun()
    
    if df.empty:
        st.info("Nenhum item aguardando aprovação.")
        return

    if "pending_order_id" in df.columns:
        df["__pending_group"] = df["pending_order_id"].fillna("").astype(str).str.strip()
        legacy_mask = df["__pending_group"] == ""
        if legacy_mask.any():
            df.loc[legacy_mask, "__pending_group"] = (
                "legacy::"
                + df.loc[legacy_mask, "SETOR2"].fillna("cliente").astype(str)
                + "::"
                + df.loc[legacy_mask, "__doc_id"].astype(str)
            )
    else:
        df["__pending_group"] = (
            "legacy::"
            + df["SETOR2"].fillna("cliente").astype(str)
            + "::"
            + df["__doc_id"].astype(str)
        )

    by_order = df.groupby("__pending_group")

    st.write(f"**Total: {len(by_order)} pedido(s) pendente(s) com {len(df)} item(ns)**")
    
    # Warning about multi-user editing
    st.info("💡 **Dica:** Clique em 🔄 Atualizar para ver alterações de outros usuários em tempo real.")
    
    st.divider()
    
    for pending_group, group_df in by_order:
        client = str(group_df.iloc[0].get("SETOR2") or "cliente")
        group_id = str(pending_group)
        group_pending_id = str(group_df.iloc[0].get("pending_order_id") or "").strip()
        # Create card container
        with st.container(border=True):
            col1, col2, col3, col4, col5 = st.columns([2, 1.5, 1, 1, 1])
            
            # Client info
            with col1:
                st.markdown(f"### 🏢 {client}")
                st.caption(f"📊 {len(group_df)} item(ns) | 💰 Total: R$ {group_df['$ Total'].sum():.2f}")
            
            # Status badge
            with col2:
                # Check if someone is editing
                lock_info = get_edit_lock_info(group_id)
                if lock_info:
                    locked_by = lock_info.get("locked_by", "unknown")
                    st.metric("Status", f"✏️ Editando: {locked_by}", delta=None)
                else:
                    st.metric("Status", "⏳ Pendente", delta=None)
            
            # Buttons
            with col3:
                if can_edit_aguardando:
                    # Check if another user is editing
                    lock_info = get_edit_lock_info(group_id)
                    is_locked_by_other = lock_info and lock_info.get("locked_by") != current_user
                    
                    if is_locked_by_other:
                        st.button("🔒", key=f"edit_{group_id}", help=f"Sendo editado por {lock_info.get('locked_by')}", disabled=True)
                    else:
                        if st.button("✏️", key=f"edit_{group_id}", help="Alterar"):
                            if st.session_state.get("edit_client") == group_id:
                                # Closing edit mode - release lock
                                release_edit_lock(group_id, current_user)
                                st.session_state.edit_client = None
                            else:
                                # Opening edit mode - try to acquire lock
                                if acquire_edit_lock(group_id, current_user):
                                    st.session_state.edit_client = group_id
                                else:
                                    st.error(f"❌ Este pedido está sendo editado por outro usuário. Aguarde.")
                                    st.rerun()

            with col4:
                if st.button(f"✅", key=f"approve_{group_id}", help="Aprovar"):
                    if firestore_enabled():
                        db = get_firestore_client()
                        batch = db.batch()
                        op_count = 0
                        max_ops = 450
                        
                        # Check for pending changes
                        has_new_items = f"new_items_{group_id}" in st.session_state and st.session_state[f"new_items_{group_id}"]
                        has_removed_items = f"removed_items_{group_id}" in st.session_state and st.session_state[f"removed_items_{group_id}"]
                        has_qty_changes = f"qtde_tracking_{group_id}" in st.session_state and st.session_state[f"qtde_tracking_{group_id}"]
                        
                        # Step 1: Apply pending changes if any
                        if has_removed_items or has_qty_changes or has_new_items:
                            # Remove items marked for deletion
                            if has_removed_items:
                                for doc_id in st.session_state[f"removed_items_{group_id}"]:
                                    doc_ref = db.collection(FIRESTORE_COLLECTION).document(doc_id)
                                    batch.delete(doc_ref)
                                    op_count += 1
                                    if op_count >= max_ops:
                                        batch.commit()
                                        batch = db.batch()
                                        op_count = 0
                            
                            # Update quantities for existing items
                            if has_qty_changes:
                                for doc_id, new_qtde in st.session_state[f"qtde_tracking_{group_id}"].items():
                                    # Only update if not marked for deletion
                                    if not has_removed_items or doc_id not in st.session_state[f"removed_items_{group_id}"]:
                                        # Find the price from the original data
                                        item_row = df[df["__doc_id"] == doc_id]
                                        if not item_row.empty:
                                            preco = float(item_row.iloc[0]["$ Unitário"])
                                            new_total = preco * new_qtde
                                            doc_ref = db.collection(FIRESTORE_COLLECTION).document(doc_id)
                                            batch.update(doc_ref, {
                                                "Qtde": int(new_qtde),
                                                "$ Total": float(new_total),
                                                "updated_at": firestore.SERVER_TIMESTAMP,
                                            })
                                            op_count += 1
                                            if op_count >= max_ops:
                                                batch.commit()
                                                batch = db.batch()
                                                op_count = 0
                            
                            # Add new items
                            if has_new_items:
                                setor = st.session_state.get("selected_setor")
                                if not setor and not group_df.empty:
                                    first_row = group_df.iloc[0]
                                    setor = {
                                        "codigo": str(first_row.get("Setor") or first_row.get("Unidade") or ""),
                                        "descricao": str(first_row.get("SETOR2") or first_row.get("Setor") or "cliente"),
                                    }
                                if setor and setor.get("codigo"):
                                    for new_item in st.session_state[f"new_items_{group_id}"]:
                                        payload = {
                                            "CòdClienteImpakto": COD_CLIENTE_IMPAKTO,
                                            "CódProImpakto": new_item["codigo"],
                                            "Item": new_item["nome"],
                                            "Qtde": int(new_item["qtde"]),
                                            "$ Unitário": float(new_item["preco"]),
                                            "$ Total": float(new_item["preco"] * new_item["qtde"]),
                                            "Unidade": setor["codigo"],
                                            "Setor": setor["codigo"],
                                            "SETOR2": setor["descricao"],
                                            "status": "aguardando",  # Will be approved in next step
                                            "pending_order_id": group_pending_id or group_id,
                                            "client_name": setor.get("descricao", "cliente"),
                                            "pdf_path": None,
                                            "created_at": firestore.SERVER_TIMESTAMP,
                                            "updated_at": firestore.SERVER_TIMESTAMP,
                                        }
                                        doc_ref = db.collection(FIRESTORE_COLLECTION).document()
                                        batch.set(doc_ref, payload)
                                        op_count += 1
                                        if op_count >= max_ops:
                                            batch.commit()
                                            batch = db.batch()
                                            op_count = 0
                            
                            # Commit pending changes
                            if op_count > 0:
                                batch.commit()
                                batch = db.batch()
                                op_count = 0
                            
                            # Clear the pending changes from session state
                            st.session_state[f"new_items_{group_id}"] = []
                            st.session_state[f"removed_items_{group_id}"] = []
                            if f"qtde_tracking_{group_id}" in st.session_state:
                                del st.session_state[f"qtde_tracking_{group_id}"]
                        
                        # Step 2: Approve only items from this pending order group
                        items_to_approve = set()
                        group_filter = group_pending_id or group_id

                        docs = db.collection(FIRESTORE_COLLECTION).where("pending_order_id", "==", group_filter).where("status", "==", "aguardando").stream()
                        for doc in docs:
                            items_to_approve.add(doc.id)

                        removed_ids = set(st.session_state.get(f"removed_items_{group_id}", []))
                        legacy_candidate_ids = [
                            str(doc_id)
                            for doc_id in group_df["__doc_id"].tolist()
                            if str(doc_id) not in removed_ids
                        ]
                        for doc_id in legacy_candidate_ids:
                            if doc_id in items_to_approve:
                                continue
                            snapshot = db.collection(FIRESTORE_COLLECTION).document(doc_id).get()
                            if snapshot.exists:
                                snapshot_data = snapshot.to_dict() or {}
                                if snapshot_data.get("status") == "aguardando":
                                    items_to_approve.add(doc_id)

                        items_to_approve = list(items_to_approve)
                        
                        # Check if there are items to approve
                        if not items_to_approve:
                            st.warning(f"⚠️ Nenhum item encontrado para aprovar. O pedido pode ter sido aprovado por outro usuário.")
                            sync_firestore_to_session()
                            st.rerun()
                            return
                        
                        # Generate a unique order number for this approval
                        order_number = get_next_order_number()
                        
                        # Approve all items with the same order_number
                        for doc_id in items_to_approve:
                            doc_ref = db.collection(FIRESTORE_COLLECTION).document(doc_id)
                            batch.update(doc_ref, {
                                "status": "pedido",
                                "order_number": order_number,
                                "approved_by": st.session_state.get("current_user", "unknown"),
                                "approved_at": firestore.SERVER_TIMESTAMP,
                                "updated_at": firestore.SERVER_TIMESTAMP,
                            })
                            op_count += 1
                            if op_count >= max_ops:
                                batch.commit()
                                batch = db.batch()
                                op_count = 0
                        
                        # Final commit
                        if op_count > 0:
                            batch.commit()
                        
                        # Release any edit lock that might exist
                        release_edit_lock(group_id, current_user)
                        
                        # Log audit trail
                        log_audit("order_approved", {
                            "order_number": order_number,
                            "client": client,
                            "items_count": len(items_to_approve),
                            "had_pending_changes": has_new_items or has_removed_items or has_qty_changes,
                        })
                        
                        sync_firestore_to_session()
                    st.success(f"✅ Pedido #{order_number} de {client} aprovado com todas as alterações! (Usuário: {st.session_state.get('current_user', 'unknown')})")
                    st.rerun()
            
            with col5:
                if st.button(f"❌", key=f"reject_{group_id}", help="Rejeitar"):
                    if firestore_enabled():
                        db = get_firestore_client()
                        for idx in group_df.index:
                            doc_id = df.loc[idx, "__doc_id"]
                            db.collection(FIRESTORE_COLLECTION).document(doc_id).delete()
                        sync_firestore_to_session()
                    st.warning(f"❌ Pedido de {client} removido!")
                    st.rerun()
            
            st.divider()
            
            # Items table
            st.write("**Items:**")
            items_display = group_df[[
                "CódProImpakto",
                "Item",
                "Qtde", 
                "$ Unitário",
                "$ Total",
                "Unidade"
            ]].copy()
            
            # Fill missing item names from product codes
            produtos_df_ref = st.session_state.excel_data.get("Produtos", pd.DataFrame()).copy()
            if not produtos_df_ref.empty:
                code_to_name_ref = {
                    str(row['productCode']): str(row['name']) for _, row in produtos_df_ref.iterrows()
                }
                items_display["Item"] = items_display.apply(
                    lambda row: code_to_name_ref.get(str(row['CódProImpakto']), row['Item']) 
                    if pd.isna(row['Item']) or str(row['Item']).strip() == "" 
                    else row['Item'],
                    axis=1
                )
            
            items_display.columns = ["📦 Código", "📝 Produto", "🔢 Qtde", "💵 Unit.", "💰 Total", "📐 Un."]
            st.dataframe(items_display, use_container_width=True, hide_index=True)
            render_pending_vs_previous_comparison(group_df, group_id)
            
            # Edit section
            is_expanded = st.session_state.get("edit_client") == group_id
            if can_edit_aguardando:
                with st.expander("✏️ Editar pedido", expanded=is_expanded):
                    st.markdown("**Editar itens do pedido:**")
                    st.caption("Altere as quantidades e remova itens conforme necessário.")

                    # Load product data
                    produtos_df = st.session_state.excel_data.get("Produtos", pd.DataFrame()).copy()
                    if not produtos_df.empty:
                        produtos_df["price"] = pd.to_numeric(
                            produtos_df.get("price"), errors="coerce"
                        ).fillna(0.0)
                        produtos_df["label"] = produtos_df.apply(
                            lambda row: f"{row['productCode']} - {row['name']}", axis=1
                        )
                        produto_options = produtos_df["label"].tolist()
                        produtos_map = {
                            row["label"]: row for _, row in produtos_df.iterrows()
                        }
                    else:
                        produtos_df = pd.DataFrame()
                        produto_options = []
                        produtos_map = {}

                    # Initialize session state for new/removed items
                    if f"new_items_{group_id}" not in st.session_state:
                        st.session_state[f"new_items_{group_id}"] = []
                    if f"removed_items_{group_id}" not in st.session_state:
                        st.session_state[f"removed_items_{group_id}"] = []

                    # Prepare data for editing
                    edit_data = []
                    for _, row in group_df.iterrows():
                        codigo = str(row["CódProImpakto"])
                        nome = row["Item"]
                        qtde = int(row["Qtde"])
                        preco = float(row["$ Unitário"])
                        doc_id = row["__doc_id"]
                        
                        edit_data.append({
                            "codigo": codigo,
                            "nome": nome,
                            "qtde": qtde,
                            "preco": preco,
                            "total": preco * qtde,
                            "doc_id": doc_id,
                            "is_new": False,
                        })
                
                    # Add new items from session state
                    for new_item in st.session_state[f"new_items_{group_id}"]:
                        edit_data.append({
                            "codigo": new_item["codigo"],
                            "nome": new_item["nome"],
                            "qtde": new_item["qtde"],
                            "preco": new_item["preco"],
                            "total": new_item["preco"] * new_item["qtde"],
                            "doc_id": None,
                            "is_new": True,
                        })
                
                    removed_ids = set(st.session_state[f"removed_items_{group_id}"])

                    # Display each item
                    edited_total = 0.0
                    qtde_dict = {}
                    
                    # Initialize qtde tracking in session state
                    if f"qtde_tracking_{group_id}" not in st.session_state:
                        st.session_state[f"qtde_tracking_{group_id}"] = {}
                    
                    for i, item in enumerate(edit_data):
                        if item["doc_id"] and item["doc_id"] in removed_ids:
                            continue
                        col1, col2, col3, col4, col5 = st.columns([2, 1, 1.2, 1, 1])
                        
                        with col1:
                            st.markdown(f"**{item['codigo']}** - {item['nome']}")
                        
                        with col2:
                            # Use session state to preserve qty changes
                            current_qty_key = f"qtde_{group_id}_{i}"
                            if item["doc_id"] and item["doc_id"] in st.session_state[f"qtde_tracking_{group_id}"]:
                                default_qty = st.session_state[f"qtde_tracking_{group_id}"][item["doc_id"]]
                            else:
                                default_qty = item["qtde"]
                            
                            new_qtde = st.number_input(
                                "Qtde",
                                min_value=1,
                                value=default_qty,
                                key=current_qty_key,
                                label_visibility="collapsed",
                            )
                            
                            # Always track the current quantity
                            if item["doc_id"]:
                                qtde_dict[item["doc_id"]] = new_qtde
                                st.session_state[f"qtde_tracking_{group_id}"][item["doc_id"]] = new_qtde
                            else:
                                # For new items, update in the list
                                for idx, new_item in enumerate(st.session_state[f"new_items_{group_id}"]):
                                    if (new_item["codigo"] == item["codigo"] and 
                                        new_item["nome"] == item["nome"] and
                                        new_item["preco"] == item["preco"]):
                                        st.session_state[f"new_items_{group_id}"][idx]["qtde"] = new_qtde
                                        break
                        
                        with col3:
                            new_total = item["preco"] * new_qtde
                            st.write(f"**R$ {new_total:.2f}**")
                            edited_total += new_total
                        
                        with col4:
                            st.write(f"Unit: R$ {item['preco']:.2f}")
                        
                        with col5:
                            if st.button("🗑️ Remover", key=f"remove_{group_id}_{i}"):
                                if item["doc_id"]:
                                    if item["doc_id"] not in st.session_state[f"removed_items_{group_id}"]:
                                        st.session_state[f"removed_items_{group_id}"].append(item["doc_id"])
                                    # Remove from tracking
                                    if item["doc_id"] in st.session_state[f"qtde_tracking_{group_id}"]:
                                        del st.session_state[f"qtde_tracking_{group_id}"][item["doc_id"]]
                                    st.rerun()
                                else:
                                    # Remove from new items
                                    idx_to_remove = None
                                    for idx, new_item in enumerate(st.session_state[f"new_items_{group_id}"]):
                                        if (new_item["codigo"] == item["codigo"] and 
                                            new_item["nome"] == item["nome"] and
                                            new_item["preco"] == item["preco"]):
                                            idx_to_remove = idx
                                            break
                                    if idx_to_remove is not None:
                                        st.session_state[f"new_items_{group_id}"].pop(idx_to_remove)
                                    st.rerun()
                
                    # Add new product button
                    st.divider()
                    st.markdown("**Adicionar novo item:**")
                    
                    col_add1, col_add2, col_add3 = st.columns([2, 1, 1])
                    
                    with col_add1:
                        if produto_options:
                            new_prod_label = st.selectbox(
                                "Produto",
                                produto_options,
                                key=f"new_prod_{group_id}",
                                label_visibility="collapsed",
                            )
                        else:
                            new_prod_label = None
                            st.warning("Nenhum produto disponível")
                    
                    with col_add2:
                        new_qtde_add = st.number_input(
                            "Qtde",
                            min_value=1,
                            value=1,
                            key=f"new_qtde_add_{group_id}",
                            label_visibility="collapsed",
                        )
                    
                    with col_add3:
                        if st.button("➕ Adicionar", key=f"add_item_{group_id}", use_container_width=True):
                            if new_prod_label and new_prod_label in produtos_map:
                                produto = produtos_map[new_prod_label]
                                new_item = {
                                    "codigo": str(produto.get("productCode")),
                                    "nome": str(produto.get("name")),
                                    "qtde": int(new_qtde_add),
                                    "preco": float(produto.get("price", 0.0)),
                                }
                                st.session_state[f"new_items_{group_id}"].append(new_item)
                                st.success(f"✅ {new_item['nome']} adicionado!")
                                st.rerun()
                
                    st.divider()
                    st.metric("Total (previa)", f"R$ {edited_total:.2f}")
                    
                    col_save, col_cancel = st.columns(2)
                    with col_save:
                        if st.button("💾 Salvar alterações", key=f"save_edit_{group_id}", use_container_width=True):
                            if not firestore_enabled():
                                st.warning("⚠️ Firestore não disponível. Alterações não foram salvas.")
                            else:
                                try:
                                    db = get_firestore_client()
                                    batch = db.batch()
                                    op_count = 0
                                    max_ops = 450
                                    changes_log = []

                                    removed_ids_list = list(st.session_state.get(f"removed_items_{group_id}", []))

                                    # Step 1: Delete removed items
                                    for doc_id in removed_ids_list:
                                        batch.delete(db.collection(FIRESTORE_COLLECTION).document(doc_id))
                                        changes_log.append(f"item removido")
                                        op_count += 1
                                        if op_count >= max_ops:
                                            batch.commit(); batch = db.batch(); op_count = 0

                                    # Step 2: Update quantities — read directly from widget state
                                    # (more reliable than qtde_dict, which depends on render order)
                                    for i, item in enumerate(edit_data):
                                        if item["is_new"] or item["doc_id"] in removed_ids_list:
                                            continue
                                        qty_key = f"qtde_{group_id}_{i}"
                                        # Read from widget state first, fall back to tracking, then original
                                        raw_qty = st.session_state.get(qty_key)
                                        if raw_qty is None:
                                            raw_qty = st.session_state.get(f"qtde_tracking_{group_id}", {}).get(item["doc_id"])
                                        if raw_qty is None:
                                            raw_qty = item["qtde"]
                                        saved_qty = max(1, int(raw_qty))
                                        new_total = item["preco"] * saved_qty
                                        batch.update(
                                            db.collection(FIRESTORE_COLLECTION).document(item["doc_id"]),
                                            {"Qtde": saved_qty, "$ Total": float(new_total), "updated_at": firestore.SERVER_TIMESTAMP},
                                        )
                                        if saved_qty != item["qtde"]:
                                            changes_log.append(f"{item['nome']}: {item['qtde']} → {saved_qty}")
                                        op_count += 1
                                        if op_count >= max_ops:
                                            batch.commit(); batch = db.batch(); op_count = 0

                                    # Step 3: Add new items
                                    new_items_list = list(st.session_state.get(f"new_items_{group_id}", []))
                                    if new_items_list:
                                        # Determine setor from multiple sources
                                        setor_codigo = ""
                                        setor_descricao = "cliente"
                                        sel = st.session_state.get("selected_setor")
                                        if sel:
                                            setor_codigo = str(sel.get("codigo", ""))
                                            setor_descricao = str(sel.get("descricao", "cliente"))
                                        if not setor_codigo and not group_df.empty:
                                            fr = group_df.iloc[0]
                                            setor_codigo = str(fr.get("Setor") or fr.get("Unidade") or "")
                                            setor_descricao = str(fr.get("SETOR2") or fr.get("client_name") or fr.get("Setor") or "cliente")
                                        if not setor_codigo:
                                            setor_codigo = str(group_pending_id or group_id)[:20]

                                        for new_item in new_items_list:
                                            payload = {
                                                "CòdClienteImpakto": COD_CLIENTE_IMPAKTO,
                                                "CódProImpakto": new_item["codigo"],
                                                "Item": new_item["nome"],
                                                "Qtde": int(new_item["qtde"]),
                                                "$ Unitário": float(new_item["preco"]),
                                                "$ Total": float(new_item["preco"] * new_item["qtde"]),
                                                "Unidade": setor_codigo,
                                                "Setor": setor_codigo,
                                                "SETOR2": setor_descricao,
                                                "status": "aguardando",
                                                "pending_order_id": group_pending_id or group_id,
                                                "client_name": setor_descricao,
                                                "pdf_path": None,
                                                "created_at": firestore.SERVER_TIMESTAMP,
                                                "updated_at": firestore.SERVER_TIMESTAMP,
                                            }
                                            batch.set(db.collection(FIRESTORE_COLLECTION).document(), payload)
                                            changes_log.append(f"adicionado: {new_item['nome']} (qty {new_item['qtde']})")
                                            op_count += 1
                                            if op_count >= max_ops:
                                                batch.commit(); batch = db.batch(); op_count = 0

                                    # Commit remaining operations
                                    if op_count:
                                        batch.commit()

                                    # Clean up ALL state for this group
                                    st.session_state.edit_client = None
                                    release_edit_lock(group_id, current_user)
                                    st.session_state[f"new_items_{group_id}"] = []
                                    st.session_state[f"removed_items_{group_id}"] = []
                                    keys_to_del = [k for k in st.session_state if k.startswith(f"qtde_{group_id}_")]
                                    for k in keys_to_del:
                                        del st.session_state[k]
                                    if f"qtde_tracking_{group_id}" in st.session_state:
                                        del st.session_state[f"qtde_tracking_{group_id}"]

                                    sync_firestore_to_session()
                                    if changes_log:
                                        st.success("✅ Salvo: " + " | ".join(changes_log))
                                    else:
                                        st.success("✅ Pedido salvo (sem alterações de quantidade).")
                                    st.rerun()
                                except Exception as _save_exc:
                                    st.error(f"❌ Erro ao salvar alterações: {_save_exc}")
                    
                    with col_cancel:
                        if st.button("❌ Cancelar", key=f"cancel_edit_{group_id}", use_container_width=True):
                            # Release edit lock
                            release_edit_lock(group_id, current_user)
                            
                            st.session_state.edit_client = None
                            st.session_state[f"new_items_{group_id}"] = []
                            st.session_state[f"removed_items_{group_id}"] = []
                            if f"qtde_tracking_{group_id}" in st.session_state:
                                del st.session_state[f"qtde_tracking_{group_id}"]
                            st.rerun()
            else:
                if is_expanded:
                    # Release lock if user leaves without proper cancel
                    release_edit_lock(group_id, current_user)
                    st.session_state.edit_client = None
                    st.rerun()
            
            st.write("")  # Spacing


def render_new_order_tab() -> None:
    data = st.session_state.excel_data
    if data is None:
        st.info("Carregue uma planilha para começar.")
        return

    if firestore_enabled():
        produtos_firestore = fetch_products_df()
        setores_firestore = fetch_setores_df()
        if not produtos_firestore.empty and not setores_firestore.empty:
            data["Produtos"] = produtos_firestore
            data["Setor"] = setores_firestore

    current_user = st.session_state.get("current_user")
    user_info = get_user_info(current_user)
    setores_df = filter_setores_for_user(data.get("Setor", pd.DataFrame()), user_info)

    st.markdown('<div class="bd-page-title">Novo Pedido</div>', unsafe_allow_html=True)

    # Card: Cliente
    st.markdown('<div class="bd-card">', unsafe_allow_html=True)
    st.markdown('<div class="bd-card-title"><h3>Selecionar Cliente</h3></div>', unsafe_allow_html=True)
    render_setor_selector(setores_df)
    has_selected_client = st.session_state.selected_setor is not None
    if not has_selected_client:
        st.caption("Selecione o cliente para habilitar a repetição do pedido.")
    if st.button("🔁 Repetir pedido anterior", use_container_width=True, disabled=not has_selected_client):
        load_latest_approved_order_into_cart()
        st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

    st.write("")  # spacing

    # Grid 2 colunas: Produto + Resumo
    col_left, col_right = st.columns([1.6, 1])

    with col_left:
        st.markdown('<div class="bd-card">', unsafe_allow_html=True)
        st.markdown('<div class="bd-card-title"><h3>Produto</h3></div>', unsafe_allow_html=True)

        produto = render_product_selector(data.get("Produtos", pd.DataFrame()))

        col_qtd, col_add = st.columns([1, 1])
        quantidade = col_qtd.number_input(
            "Quantidade",
            min_value=1,
            max_value=9999,
            value=1,
            step=1,
            label_visibility="collapsed",
        )
        can_add = produto is not None and st.session_state.selected_setor is not None

        with col_add:
            st.markdown('<div class="bd-primary">', unsafe_allow_html=True)
            if st.button("Adicionar ao carrinho", disabled=not can_add, use_container_width=True):
                add_item_to_cart(produto, int(quantidade))
                st.success("Produto adicionado ao carrinho.")
            st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('</div>', unsafe_allow_html=True)

        st.write("")
        st.markdown('<div class="bd-card">', unsafe_allow_html=True)
        st.markdown('<div class="bd-card-title"><h3>Itens do pedido</h3></div>', unsafe_allow_html=True)
        render_cart(compact=True)
        st.markdown('</div>', unsafe_allow_html=True)

    with col_right:
        st.markdown('<div class="bd-card">', unsafe_allow_html=True)
        st.markdown('<div class="bd-card-title"><h3>Ações</h3></div>', unsafe_allow_html=True)
        render_previa_download()
        render_save_buttons()
        st.markdown('</div>', unsafe_allow_html=True)


def render_approved_orders_tab() -> None:
    """Exibe pedidos aprovados agrupados por número de pedido."""
    st.subheader("📋 Pedidos Aprovados - Histórico")
    
    try:
        db = get_firestore_client()
        
        # Fetch approved orders
        try:
            approved_orders = db.collection(FIRESTORE_COLLECTION)\
                .where("status", "==", "pedido")\
                .order_by("approved_at", direction=firestore.Query.DESCENDING)\
                .stream()

            orders_list = []
            for doc in approved_orders:
                order_data = doc.to_dict()
                order_data["__doc_id"] = doc.id
                orders_list.append(order_data)
        except Exception as exc:
            st.warning(
                "⚠️ Índice não encontrado no Firestore. "
                "Carregando sem ordenação (fallback local)."
            )
            approved_orders = db.collection(FIRESTORE_COLLECTION)\
                .where("status", "==", "pedido")\
                .stream()

            orders_list = []
            for doc in approved_orders:
                order_data = doc.to_dict()
                order_data["__doc_id"] = doc.id
                orders_list.append(order_data)

            def sort_key(item: Dict) -> datetime:
                value = item.get("approved_at")
                if hasattr(value, "to_datetime"):
                    return value.to_datetime()
                if isinstance(value, datetime):
                    return value
                return datetime.min

            orders_list.sort(key=sort_key, reverse=True)
        
        if not orders_list:
            st.info("ℹ️ Nenhum pedido aprovado ainda.")
            return
        
        # Convert to DataFrame and group by order_number
        orders_df = pd.DataFrame(orders_list)
        
        # For orders without order_number (legacy), use __doc_id as fallback
        if "order_number" not in orders_df.columns:
            orders_df["order_number"] = orders_df["__doc_id"]
        else:
            orders_df["order_number"] = orders_df["order_number"].fillna(orders_df["__doc_id"])
        
        # Group by client first
        grouped_by_client = orders_df.groupby("SETOR2")
        
        st.write(f"**Total: {len(grouped_by_client)} cliente(s) com {len(orders_df.groupby('order_number'))} pedido(s)**")
        st.divider()
        
        for client, client_df in grouped_by_client:
            with st.container(border=True):
                # Group by order_number within each client
                grouped_by_order = client_df.groupby("order_number")
                
                col1, col2, col3 = st.columns([2, 1, 1])
                col1.markdown(f"### 🏢 {client}")
                col2.metric("Pedidos", len(grouped_by_order))
                col3.metric("Total", f"R$ {client_df['$ Total'].sum():.2f}")
                
                st.divider()
                
                # Display each order (grouped by order_number)
                for order_num, order_items in grouped_by_order:
                    # Get order info from first item
                    first_item = order_items.iloc[0]
                    approved_at = first_item.get('approved_at', 'N/A')
                    if hasattr(approved_at, 'to_datetime'):
                        approved_at = approved_at.to_datetime().strftime('%Y-%m-%d %H:%M:%S')
                    approved_by = first_item.get('approved_by', 'unknown')
                    created_at = first_item.get('created_at', 'N/A')
                    if hasattr(created_at, 'to_datetime'):
                        created_at = created_at.to_datetime().strftime('%Y-%m-%d %H:%M:%S')
                    
                    order_total = order_items['$ Total'].sum()
                    items_count = len(order_items)
                    
                    with st.expander(f"📦 Pedido #{order_num} - {items_count} item(ns) - R$ {order_total:.2f}"):
                        col_info, col_user = st.columns([3, 1])
                        
                        with col_info:
                            st.write("**Informações:**")
                            st.write(f"- 📅 Criado em: {created_at}")
                            st.write(f"- ✅ Aprovado em: {approved_at}")
                            st.write(f"- 📦 Itens: {items_count}")
                            st.write(f"- 💰 Total: R$ {order_total:.2f}")
                        
                        with col_user:
                            st.write("**Aprovado por:**")
                            st.write(f"👤 {approved_by}")
                        
                        st.divider()
                        
                        # Display items table
                        st.write("**Itens do Pedido:**")
                        items_display = order_items[[
                            "CódProImpakto",
                            "Item",
                            "Qtde", 
                            "$ Unitário",
                            "$ Total",
                            "Unidade"
                        ]].copy()
                        items_display.columns = ["📦 Código", "📝 Produto", "🔢 Qtde", "💵 Unit.", "💰 Total", "📐 Un."]
                        st.dataframe(items_display, use_container_width=True, hide_index=True)
                        
                        # Admin delete button
                        if st.session_state.get("current_user") == "admin":
                            st.divider()
                            col_delete = st.columns([1, 3])
                            with col_delete[0]:
                                confirm_key = f"confirm_delete_order_{order_num}"
                                confirmed = st.checkbox("Confirmar exclusão", key=confirm_key)
                                if st.button(
                                    "🗑️ Deletar Pedido",
                                    key=f"delete_order_{order_num}",
                                    help="Somente admin - deleta todos os itens deste pedido",
                                    disabled=not confirmed,
                                ):
                                    # Delete all items with this order_number
                                    for doc_id in order_items["__doc_id"]:
                                        db.collection(FIRESTORE_COLLECTION).document(doc_id).delete()
                                    st.success(f"✅ Pedido #{order_num} deletado com sucesso!")
                                    sync_firestore_to_session()
                                    st.rerun()
                
                st.write("")  # Spacing
    
    except Exception as e:
        st.error(f"❌ Erro ao carregar pedidos aprovados: {e}")


def render_history_tab() -> None:
    """Display order history with audit trail."""
    st.subheader("📚 Histórico de Pedidos")
    
    try:
        db = get_firestore_client()
        
        # Fetch all history records
        try:
            history_orders = db.collection("historico_pedidos")\
                .order_by("export_date", direction=firestore.Query.DESCENDING)\
                .stream()
            
            orders_list = []
            for doc in history_orders:
                order_data = doc.to_dict()
                order_data["__doc_id"] = doc.id
                orders_list.append(order_data)
        except Exception as exc:
            st.warning("⚠️ Índice não encontrado. Carregando sem ordenação.")
            history_orders = db.collection("historico_pedidos").stream()
            
            orders_list = []
            for doc in history_orders:
                order_data = doc.to_dict()
                order_data["__doc_id"] = doc.id
                orders_list.append(order_data)
            
            def sort_key(item: Dict) -> datetime:
                value = item.get("export_date")
                if hasattr(value, "to_datetime"):
                    return value.to_datetime()
                if isinstance(value, datetime):
                    return value
                return datetime.min
            
            orders_list.sort(key=sort_key, reverse=True)
        
        if not orders_list:
            st.info("ℹ️ Nenhum pedido no histórico ainda.")
            return
        
        # Filters and grouping
        orders_df = pd.DataFrame(orders_list)
        if "Setor" not in orders_df.columns:
            orders_df["Setor"] = "N/A"
        if "client_name" not in orders_df.columns:
            orders_df["client_name"] = ""

        orders_df["Setor"] = orders_df["Setor"].fillna("N/A").astype(str)
        orders_df["client_name"] = orders_df["client_name"].fillna("").astype(str)
        orders_df["created_datetime"] = orders_df["created_at"].apply(coerce_datetime)
        orders_df["created_date"] = orders_df["created_datetime"].apply(
            lambda value: value.date() if isinstance(value, datetime) else None
        )
        orders_df["created_date_iso"] = orders_df["created_at"].apply(extract_date_iso)

        available_dates = [value for value in orders_df["created_date"].tolist() if isinstance(value, date)]
        default_start = min(available_dates) if available_dates else date.today()
        default_end = max(available_dates) if available_dates else date.today()

        filter_col1, filter_col2 = st.columns([1, 1])
        with filter_col1:
            use_exact_filter = st.checkbox("🔎 Filtrar por data exata", key="history_use_exact_date_filter")
        with filter_col2:
            use_range_filter = st.checkbox("📅 Filtrar por intervalo", key="history_use_range_date_filter")

        created_exact_date = None
        if use_exact_filter:
            created_exact_date = st.date_input(
                "Data de criação",
                value=default_end,
                key="history_created_date_exact_picker",
                format="YYYY-MM-DD",
            )

        created_from_date = None
        created_to_date = None
        if use_range_filter:
            range_col1, range_col2 = st.columns([1, 1])
            with range_col1:
                created_from_date = st.date_input(
                    "De",
                    value=default_start,
                    key="history_created_date_from_picker",
                    format="YYYY-MM-DD",
                )
            with range_col2:
                created_to_date = st.date_input(
                    "Até",
                    value=default_end,
                    key="history_created_date_to_picker",
                    format="YYYY-MM-DD",
                )

        if created_exact_date:
            orders_df = orders_df[orders_df["created_date"] == created_exact_date]

        if created_from_date and created_to_date and created_from_date > created_to_date:
            st.warning("⚠️ A data inicial não pode ser maior que a data final.")
            return

        if created_from_date:
            orders_df = orders_df[orders_df["created_date"].notna()]
            orders_df = orders_df[orders_df["created_date"] >= created_from_date]

        if created_to_date:
            orders_df = orders_df[orders_df["created_date"].notna()]
            orders_df = orders_df[orders_df["created_date"] <= created_to_date]

        if orders_df.empty:
            st.info("ℹ️ Nenhum pedido encontrado para o filtro informado.")
            return

        filtered_doc_ids = orders_df["__doc_id"].astype(str).tolist()
        select_col1, select_col2, select_col3 = st.columns([2, 1, 1])
        with select_col1:
            mark_all = st.checkbox("Selecionar todos os pedidos filtrados", key="history_mark_all_checkbox")
        with select_col2:
            if st.button("Aplicar seleção", key="history_apply_select_all", use_container_width=True):
                for doc_id in filtered_doc_ids:
                    st.session_state[f"history_selected_{doc_id}"] = mark_all
                st.rerun()
        with select_col3:
            if st.button("Limpar seleção", key="history_clear_selection", use_container_width=True):
                for doc_id in filtered_doc_ids:
                    st.session_state[f"history_selected_{doc_id}"] = False
                st.rerun()

        selected_orders = [
            order for order in orders_df.to_dict("records")
            if st.session_state.get(f"history_selected_{str(order.get('__doc_id'))}", False)
        ]

        action_col1, action_col2 = st.columns([1, 2])
        with action_col1:
            st.metric("Selecionados", len(selected_orders))
        with action_col2:
            if selected_orders:
                batch_pdf = generate_history_batch_pdf(selected_orders)
                st.download_button(
                    "🖨️ Imprimir selecionados (1 pedido por página)",
                    data=batch_pdf.getvalue(),
                    file_name=f"HISTORICO_Selecionados_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                    mime="application/pdf",
                    key="history_download_selected_orders_pdf",
                    use_container_width=True,
                )
            else:
                st.info("Selecione pedidos (ou clique em Aplicar seleção) para liberar a impressão em lote.")

        grouped = orders_df.groupby(["Setor", "client_name"], dropna=False)
        
        for (setor_code, client_name), group_df in grouped:
            setor_label = str(setor_code)
            client_label = client_name.strip() if isinstance(client_name, str) else ""
            if not client_label:
                client_label = "Sem nome"
            with st.container(border=True):
                col1, col2, col3 = st.columns([2, 1, 1])
                col1.markdown(f"**👤 Cliente: {client_label}**")
                col1.caption(f"🏷️ Setor: {setor_label}")
                col2.metric("Pedidos", len(group_df))
                col3.metric("Total", f"R$ {group_df['$ Total'].sum():.2f}")
                
                # Display each history order
                for idx, order in group_df.iterrows():
                    doc_id = str(order.get('__doc_id'))
                    created_at = format_datetime_display(order.get('created_at', 'N/A'))
                    export_date = format_datetime_display(order.get('export_date', 'N/A'))
                    total = parse_float(order.get('$ Total', 0))
                    
                    with st.expander(
                        f"📦 Pedido - Cliente: {client_label} - Setor: {setor_label} - Exportado em {export_date}"
                    ):
                        st.checkbox(
                            "Selecionar este pedido para impressão em lote",
                            key=f"history_selected_{doc_id}",
                        )

                        col_info, col_dates = st.columns([2, 1])
                        
                        with col_info:
                            st.write("**Informações do Pedido:**")
                            st.write(f"- Cliente: {client_label}")
                            st.write(f"- Setor: {setor_label}")
                            st.write(f"- Criado em: {created_at}")
                            st.write(f"- Total: R$ {total:.2f}")
                        
                        with col_dates:
                            st.write("**Datas:**")
                            st.write(f"- Aprovado em: {format_datetime_display(order.get('approved_at', 'N/A'))}")
                            st.write(f"- Exportado em: {export_date}")

                            pdf_buffer = generate_history_order_pdf(order, client_label, setor_label)
                            safe_setor = str(setor_label).replace(" ", "_").replace("/", "-")
                            safe_order = str(order.get("order_number") or order.get("__doc_id") or "sem_numero")
                            st.download_button(
                                "🖨️ Baixar PDF para imprimir",
                                data=pdf_buffer.getvalue(),
                                file_name=f"HISTORICO_Pedido_{safe_setor}_{safe_order}.pdf",
                                mime="application/pdf",
                                key=f"download_history_pdf_{doc_id}",
                                use_container_width=True,
                            )
                        
                        st.divider()
                        
                        # Display items
                        st.write("**Itens do Pedido:**")
                        items = order.get('items', [])
                        if items:
                            items_df = pd.DataFrame(items)
                            items_df.columns = ["📦 Código", "📝 Produto", "🔢 Qtde", "💵 Unit.", "💰 Total", "📐 Un."]
                            st.dataframe(items_df, use_container_width=True, hide_index=True)
                        else:
                            st.write("- Nenhum item registrado")
                        
                        st.divider()
                        
                        # Display audit trail with double-click functionality
                        st.write("**Histórico de Alterações (durante Aguardando Aprovação):**")
                        
                        audit_entries = order.get('audit_trail', [])
                        if audit_entries:
                            for i, entry in enumerate(audit_entries):
                                action = entry.get('action', 'N/A')
                                timestamp = entry.get('timestamp', 'N/A')
                                details = entry.get('details', {})
                                
                                # Display audit entry
                                col1, col2 = st.columns([3, 1])
                                with col1:
                                    st.write(f"- **{action}** em {timestamp}")
                                    if action == "item_added":
                                        product_name = details.get('product_name', 'Produto desconhecido')
                                        qtde = details.get('quantity', 0)
                                        st.caption(f"  ➕ Adicionado: {product_name} (Qtde: {qtde})")
                                    elif action == "item_removed":
                                        product_name = details.get('product_name', 'Produto desconhecido')
                                        qtde = details.get('quantity', 0)
                                        st.caption(f"  ➖ Removido: {product_name} (Qtde: {qtde})")
                                    elif action == "quantity_changed":
                                        product_name = details.get('product_name', 'Produto desconhecido')
                                        old_qty = details.get('old_quantity', 0)
                                        new_qty = details.get('new_quantity', 0)
                                        st.caption(f"  🔄 {product_name}: {old_qty} → {new_qty}")
                        else:
                            st.write("- Nenhuma alteração registrada")

                        if st.session_state.get("current_user") == "admin":
                            st.divider()
                            confirm_key = f"confirm_delete_history_{order.get('__doc_id')}"
                            confirmed = st.checkbox("Confirmar", key=confirm_key)
                            if st.button(
                                "🗑️ Apagar do histórico",
                                key=f"delete_history_{order.get('__doc_id')}",
                                help="Remove o pedido do histórico",
                                disabled=not confirmed,
                            ):
                                db.collection("historico_pedidos").document(order.get("__doc_id")).delete()
                                st.success("✅ Pedido removido do histórico!")
                                st.rerun()

    except Exception as e:
        st.error(f"❌ Erro ao carregar histórico: {e}")


def render_catalog_editor_tab() -> None:
    st.subheader("✏️ Editar Catálogo")

    if not can_edit_catalog():
        st.warning("⚠️ Você não tem permissão para editar o catálogo.")
        return

    tab_prod, tab_setor, tab_hist_prod, tab_hist_setor = st.tabs([
        "📦 Produtos", "🏢 Setores", "📋 Histórico Produtos", "📋 Histórico Setores"
    ])

    # ── Produtos ──────────────────────────────────────────────────────────────
    with tab_prod:
        st.caption(
            "Edite preços e nomes diretamente na tabela. "
            "Use a última linha em branco para adicionar novos produtos. "
            "Clique em **Salvar Produtos** para enviar ao Firestore."
        )

        raw_prod = fetch_products_df()
        prod_df = normalize_products_df(raw_prod)
        prod_df = prod_df.drop(columns=["__doc_id"], errors="ignore")

        edited_prod = st.data_editor(
            prod_df,
            num_rows="dynamic",
            use_container_width=True,
            hide_index=True,
            column_config={
                "productCode": st.column_config.TextColumn("Código", required=True),
                "name": st.column_config.TextColumn("Nome do Produto", width="large"),
                "price": st.column_config.NumberColumn(
                    "Preço (R$)", format="R$ %.2f", min_value=0.0, step=0.01
                ),
            },
            key="catalog_editor_produtos",
        )

        col_save, col_reset = st.columns([2, 1])
        with col_save:
            if st.button("💾 Salvar Produtos", type="primary", use_container_width=True):
                new_prod = normalize_products_df(edited_prod)
                if new_prod.empty:
                    st.error("Nenhum produto válido para salvar.")
                else:
                    changes = compute_product_changes(prod_df, new_prod)
                    if firestore_enabled():
                        upsert_collection_from_df(FIRESTORE_PRODUCTS_COLLECTION, new_prod, "productCode")
                        fetch_products_df.clear()
                        sync_firestore_reference_data()
                    else:
                        if st.session_state.excel_data is None:
                            st.session_state.excel_data = {}
                        st.session_state.excel_data["Produtos"] = new_prod
                    save_product_history(changes, "editor_catalogo")
                    log_audit(
                        "catalog_products_saved",
                        {"items_changed": int(len(changes)), "total": int(len(new_prod))},
                    )
                    st.success(f"✅ {len(new_prod)} produto(s) salvos. {len(changes)} alteração(ões) registrada(s).")
                    st.rerun()
        with col_reset:
            if st.button("🔄 Recarregar do Firestore", use_container_width=True):
                fetch_products_df.clear()
                sync_firestore_reference_data()
                st.rerun()

    # ── Setores ───────────────────────────────────────────────────────────────
    with tab_setor:
        st.caption(
            "Edite o código e descrição dos setores. "
            "Use a última linha em branco para adicionar novos setores. "
            "Clique em **Salvar Setores** para enviar ao Firestore."
        )

        raw_set = fetch_setores_df()
        set_df = raw_set.drop(columns=["__doc_id"], errors="ignore").copy()
        for col in ["CódUnidade", "items__description"]:
            if col not in set_df.columns:
                set_df[col] = ""
        set_df = set_df[["CódUnidade", "items__description"]]
        set_df["CódUnidade"] = set_df["CódUnidade"].astype(str).str.strip()
        set_df["items__description"] = set_df["items__description"].astype(str).str.strip()
        set_df = set_df[set_df["CódUnidade"].str.len() > 0].sort_values("CódUnidade").reset_index(drop=True)

        edited_set = st.data_editor(
            set_df,
            num_rows="dynamic",
            use_container_width=True,
            hide_index=True,
            column_config={
                "CódUnidade": st.column_config.TextColumn("Código do Setor", required=True),
                "items__description": st.column_config.TextColumn("Descrição do Setor", width="large"),
            },
            key="catalog_editor_setores",
        )

        col_save2, col_reset2 = st.columns([2, 1])
        with col_save2:
            if st.button("💾 Salvar Setores", type="primary", use_container_width=True):
                new_set = edited_set.copy()
                new_set["CódUnidade"] = new_set["CódUnidade"].astype(str).str.strip()
                new_set = new_set[new_set["CódUnidade"].str.len() > 0].drop_duplicates(subset=["CódUnidade"], keep="last")
                if new_set.empty:
                    st.error("Nenhum setor válido para salvar.")
                else:
                    save_setores_history(set_df, new_set)
                    if firestore_enabled():
                        upsert_collection_from_df(FIRESTORE_SETORES_COLLECTION, new_set, "CódUnidade")
                        fetch_setores_df.clear()
                        sync_firestore_reference_data()
                    else:
                        if st.session_state.excel_data is None:
                            st.session_state.excel_data = {}
                        st.session_state.excel_data["Setor"] = new_set
                    log_audit("catalog_setores_saved", {"total": int(len(new_set))})
                    st.success(f"✅ {len(new_set)} setor(es) salvos com sucesso.")
                    st.rerun()
        with col_reset2:
            if st.button("🔄 Recarregar do Firestore", use_container_width=True, key="reload_setores"):
                fetch_setores_df.clear()
                sync_firestore_reference_data()
                st.rerun()

    # ── Histórico de Produtos ─────────────────────────────────────────────────
    with tab_hist_prod:
        st.caption("Registro de todas as alterações de produtos com data e responsável.")
        if st.button("🔄 Atualizar histórico", key="refresh_hist_prod"):
            _fetch_product_history_firestore.clear()
            st.rerun()
        hist_prod = fetch_product_history(limit=500)
        if hist_prod.empty:
            st.info("Nenhuma alteração de produto registrada ainda.")
        else:
            disp = hist_prod.copy()
            disp["Data"] = disp.get("changed_at", pd.Series()).apply(
                lambda v: coerce_datetime(v).strftime("%d/%m/%Y %H:%M") if coerce_datetime(v) else str(v)
            )
            disp["Preço Anterior"] = pd.to_numeric(disp.get("old_price", 0), errors="coerce").fillna(0).apply(format_currency)
            disp["Preço Novo"] = pd.to_numeric(disp.get("new_price", 0), errors="coerce").fillna(0).apply(format_currency)
            disp["Diferença"] = pd.to_numeric(disp.get("diff_price", 0), errors="coerce").fillna(0).apply(format_currency)
            disp = disp.rename(columns={
                "productCode": "Código",
                "name": "Produto",
                "change_type": "Tipo",
                "changed_by": "Alterado por",
                "source": "Origem",
            })
            cols = [c for c in ["Data", "Alterado por", "Código", "Produto", "Tipo", "Preço Anterior", "Preço Novo", "Diferença", "Origem"] if c in disp.columns]
            st.dataframe(disp[cols], use_container_width=True, hide_index=True)

    # ── Histórico de Setores ──────────────────────────────────────────────────
    with tab_hist_setor:
        st.caption("Registro de todas as alterações de setores com data e responsável.")
        if st.button("🔄 Atualizar histórico", key="refresh_hist_setor"):
            _fetch_setores_history_firestore.clear()
            st.rerun()
        hist_set = fetch_setores_history(limit=200)
        if hist_set.empty:
            st.info("Nenhuma alteração de setor registrada ainda.")
        else:
            disp2 = hist_set.copy()
            disp2["Data"] = disp2.get("changed_at", pd.Series()).apply(
                lambda v: coerce_datetime(v).strftime("%d/%m/%Y %H:%M") if coerce_datetime(v) else str(v)
            )
            disp2 = disp2.rename(columns={
                "codigo": "Código",
                "old_desc": "Descrição Anterior",
                "new_desc": "Descrição Nova",
                "change_type": "Tipo",
                "changed_by": "Alterado por",
            })
            cols2 = [c for c in ["Data", "Alterado por", "Código", "Tipo", "Descrição Anterior", "Descrição Nova"] if c in disp2.columns]
            st.dataframe(disp2[cols2], use_container_width=True, hide_index=True)


def render_product_update_tab() -> None:
    st.subheader("📦 Atualização de Produtos")

    if st.session_state.get("current_user") != "admin":
        st.warning("⚠️ Apenas admins podem atualizar produtos.")
        return

    current_products = st.session_state.excel_data.get("Produtos", pd.DataFrame())
    current_products = normalize_products_df(current_products)

    st.caption("Envie uma planilha para comparar preços com a base atual antes de atualizar.")
    upload = st.file_uploader(
        "Planilha para atualização de produtos (.xlsx)",
        type=["xlsx", "xls"],
        key="product_update_upload",
    )

    source_label = st.session_state.get("product_update_source", "manual")
    if upload is not None:
        try:
            upload_data = load_excel_data(upload.getvalue())
            incoming_products = normalize_products_df(upload_data.get("Produtos", pd.DataFrame()))
            if incoming_products.empty:
                st.error("A planilha enviada não possui dados válidos na aba 'Produtos'.")
            else:
                changes = compute_product_changes(current_products, incoming_products)
                st.session_state.product_update_changes = changes
                st.session_state.product_update_df = incoming_products
                source_label = f"upload ({upload.name})"
                st.session_state.product_update_source = source_label

                col_a, col_b, col_c = st.columns(3)
                col_a.metric("Produtos atuais", len(current_products))
                col_b.metric("Produtos na planilha", len(incoming_products))
                col_c.metric("Alterações detectadas", len(changes))

                if changes.empty:
                    st.info("Nenhuma alteração de produto/preço foi detectada.")
                else:
                    preview = changes.copy()
                    preview["Preço Anterior"] = preview["old_price"].apply(format_currency)
                    preview["Preço Novo"] = preview["new_price"].apply(format_currency)
                    preview["Diferença"] = preview["diff_price"].apply(format_currency)
                    preview = preview.rename(
                        columns={
                            "productCode": "Código",
                            "name": "Produto",
                            "change_type": "Tipo",
                        }
                    )
                    st.markdown("**Pré-visualização das alterações**")
                    st.dataframe(
                        preview[["Código", "Produto", "Tipo", "Preço Anterior", "Preço Novo", "Diferença"]],
                        use_container_width=True,
                        hide_index=True,
                    )
        except Exception as exc:
            st.error(f"❌ Erro ao analisar planilha: {exc}")

    can_apply = not st.session_state.get("product_update_df", pd.DataFrame()).empty
    if st.button("✅ Aplicar atualização de produtos", use_container_width=True, disabled=not can_apply):
        incoming_products = st.session_state.get("product_update_df", pd.DataFrame())
        changes = st.session_state.get("product_update_changes", pd.DataFrame())
        if incoming_products.empty:
            st.error("Nenhuma planilha preparada para atualização.")
        else:
            source_label = st.session_state.get("product_update_source", "manual")
            if firestore_enabled():
                upsert_collection_from_df(FIRESTORE_PRODUCTS_COLLECTION, incoming_products, "productCode")
                fetch_products_df.clear()
                sync_firestore_reference_data()
            else:
                st.session_state.excel_data["Produtos"] = incoming_products

            save_product_history(changes, source_label)
            log_audit(
                "products_updated",
                {
                    "source": source_label,
                    "items_changed": int(len(changes)),
                    "total_products": int(len(incoming_products)),
                },
            )
            st.success(f"✅ Produtos atualizados com sucesso. Alterações registradas: {len(changes)}")
            st.session_state.product_update_changes = pd.DataFrame()
            st.session_state.product_update_df = pd.DataFrame()
            st.session_state.product_update_source = "manual"
            st.rerun()

    st.divider()
    st.markdown("**Histórico de alterações de produto**")
    history_df = fetch_product_history(limit=500)
    if history_df.empty:
        st.info("Nenhum histórico de alteração de produto encontrado.")
        return

    display_df = history_df.copy()
    display_df["Preço Anterior"] = pd.to_numeric(display_df.get("old_price", 0.0), errors="coerce").fillna(0.0).apply(format_currency)
    display_df["Preço Novo"] = pd.to_numeric(display_df.get("new_price", 0.0), errors="coerce").fillna(0.0).apply(format_currency)
    display_df["Diferença"] = pd.to_numeric(display_df.get("diff_price", 0.0), errors="coerce").fillna(0.0).apply(format_currency)
    if "changed_at" in display_df.columns:
        display_df["Data"] = display_df["changed_at"].apply(
            lambda value: coerce_datetime(value).strftime("%d/%m/%Y %H:%M") if coerce_datetime(value) else str(value)
        )
    else:
        display_df["Data"] = "-"

    display_df = display_df.rename(
        columns={
            "productCode": "Código",
            "name": "Produto",
            "change_type": "Tipo",
            "changed_by": "Alterado por",
            "source": "Origem",
        }
    )
    st.dataframe(
        display_df[["Data", "Código", "Produto", "Tipo", "Preço Anterior", "Preço Novo", "Diferença", "Alterado por", "Origem"]],
        use_container_width=True,
        hide_index=True,
    )

    st.divider()
    st.markdown("**Relatório de aumento de preços**")

    report_df = history_df.copy()
    report_df["diff_price_num"] = pd.to_numeric(report_df.get("diff_price", 0.0), errors="coerce").fillna(0.0)
    report_df = report_df[
        (report_df.get("change_type", "") == "Preço Alterado")
        & (report_df["diff_price_num"] > 0)
    ].copy()

    if report_df.empty:
        st.info("Nenhum aumento de preço registrado até o momento.")
        return

    report_df["changed_at_dt"] = report_df.get("changed_at").apply(coerce_datetime)
    report_df = report_df[report_df["changed_at_dt"].notna()].copy()

    if report_df.empty:
        st.info("Não foi possível converter datas do histórico para montar o relatório.")
        return

    min_date = report_df["changed_at_dt"].min().date()
    max_date = report_df["changed_at_dt"].max().date()

    c1, c2 = st.columns(2)
    with c1:
        start_date = st.date_input("Data inicial", value=min_date, min_value=min_date, max_value=max_date, key="price_report_start")
    with c2:
        end_date = st.date_input("Data final", value=max_date, min_value=min_date, max_value=max_date, key="price_report_end")

    if start_date > end_date:
        st.warning("A data inicial não pode ser maior que a data final.")
        return

    filtered = report_df[
        (report_df["changed_at_dt"].dt.date >= start_date)
        & (report_df["changed_at_dt"].dt.date <= end_date)
    ].copy()

    if filtered.empty:
        st.info("Não há aumentos de preço no período selecionado.")
        return

    filtered["Data"] = filtered["changed_at_dt"].dt.strftime("%d/%m/%Y %H:%M")
    filtered["Preço Anterior"] = pd.to_numeric(filtered.get("old_price", 0.0), errors="coerce").fillna(0.0).apply(format_currency)
    filtered["Preço Novo"] = pd.to_numeric(filtered.get("new_price", 0.0), errors="coerce").fillna(0.0).apply(format_currency)
    filtered["Aumento"] = filtered["diff_price_num"].apply(format_currency)

    total_increases = float(filtered["diff_price_num"].sum())
    products_changed = int(filtered["productCode"].nunique()) if "productCode" in filtered.columns else 0
    events_count = int(len(filtered))

    m1, m2, m3 = st.columns(3)
    m1.metric("Eventos de aumento", events_count)
    m2.metric("Produtos afetados", products_changed)
    m3.metric("Soma dos aumentos", format_currency(total_increases))

    filtered = filtered.rename(
        columns={
            "productCode": "Código",
            "name": "Produto",
            "changed_by": "Alterado por",
            "source": "Origem",
        }
    )

    st.dataframe(
        filtered[["Data", "Código", "Produto", "Preço Anterior", "Preço Novo", "Aumento", "Alterado por", "Origem"]],
        use_container_width=True,
        hide_index=True,
    )


def render_user_management_tab() -> None:
    """Admin interface for user management."""
    st.subheader("👥 Gerenciamento de Usuários")
    
    # Check if current user is admin
    if st.session_state.get("current_user") != "admin":
        st.warning("⚠️ Apenas admins podem gerenciar usuários.")
        return
    
    # Tab: Create, List, Edit
    tab_create, tab_list, tab_edit = st.tabs(["➕ Criar Usuário", "📋 Listar Usuários", "✏️ Editar Usuário"])
    
    # TAB 1: Create User
    with tab_create:
        st.markdown("#### Criar Novo Usuário")
        col1, col2 = st.columns([1, 1])
        
        with col1:
            new_username = st.text_input("👤 Nome de usuário")
        with col2:
            new_role = st.selectbox("📌 Função", ["user", "admin", "supervisora", "catalog_editor"])
        
        new_password = st.text_input("🔑 Senha", type="password")
        confirm_password = st.text_input("🔑 Confirmar Senha", type="password")
        
        # Setor assignment for supervisora
        assigned_setores = None
        if new_role == "supervisora":
            st.markdown("---")
            st.markdown("**Setores Vinculados:**")
            setores_df = st.session_state.excel_data.get("Setor", pd.DataFrame())
            if not setores_df.empty and "CódUnidade" in setores_df.columns and "items__description" in setores_df.columns:
                # Create label column like in render_setor_selector
                setores_temp = setores_df.dropna(subset=["CódUnidade", "items__description"]).copy()
                setores_temp["codigo"] = setores_temp["CódUnidade"].apply(parse_int)
                setores_temp["descricao"] = setores_temp["items__description"].astype(str)
                setores_temp["label"] = setores_temp.apply(
                    lambda row: f"{row['codigo']} - {row['descricao']}", axis=1
                )
                available_setores = setores_temp["label"].tolist()
                assigned_setores = st.multiselect(
                    "Selecione os setores que esta supervisora pode acessar",
                    options=available_setores,
                    help="Supervisora verá apenas estes setores"
                )
            else:
                st.warning("⚠️ Carregue a planilha primeiro para selecionar setores.")
        
        if st.button("✅ Criar Usuário", use_container_width=True):
            if not new_username or not new_password:
                st.error("❌ Usuário e senha são obrigatórios.")
            elif new_password != confirm_password:
                st.error("❌ As senhas não correspondem.")
            elif new_role == "supervisora" and not assigned_setores:
                st.error("❌ Selecione pelo menos um setor para a supervisora.")
            else:
                if create_user(new_username, new_password, new_role, assigned_setores):
                    st.balloons()
    
    # TAB 2: List Users
    with tab_list:
        st.markdown("#### Lista de Usuários")
        users = get_users_list()
        
        if not users:
            st.info("ℹ️ Nenhum usuário cadastrado.")
        else:
            users_df = pd.DataFrame(users)
            users_df = users_df[["username", "role", "created_at", "updated_at"]]
            st.dataframe(users_df, use_container_width=True, hide_index=True)
    
    # TAB 3: Edit User
    with tab_edit:
        st.markdown("#### Editar Usuário")
        users = get_users_list()
        
        if not users:
            st.info("ℹ️ Nenhum usuário para editar.")
        else:
            usernames = [u["username"] for u in users]
            selected_user = st.selectbox("👤 Selecione usuário", usernames)

            if st.session_state.get("user_action_message"):
                st.success(st.session_state.pop("user_action_message"))
            if st.session_state.get("user_action_error"):
                st.error(st.session_state.pop("user_action_error"))
            
            col1, col2, col3 = st.columns([1, 1, 1])

            with col1:
                st.markdown("**Redefinir Senha**")
                new_pwd = st.text_input("Nova Senha", type="password", key="reset_pwd")
                confirm_pwd = st.text_input("Confirmar Senha", type="password", key="confirm_pwd")
                if st.button("✅ Atualizar Senha", use_container_width=True):
                    if not new_pwd:
                        st.error("❌ Informe a nova senha.")
                    elif new_pwd != confirm_pwd:
                        st.error("❌ Senhas não correspondem.")
                    else:
                        if update_user_password(selected_user, new_pwd):
                            st.session_state["user_action_message"] = "✅ Senha atualizada!"
                            st.rerun()

            with col2:
                st.markdown("**Deletar Usuário**")
                confirm_delete = st.checkbox(
                    f"Confirmo excluir '{selected_user}'",
                    key=f"confirm_delete_user_{selected_user}",
                )
                if st.button(
                    "🗑️ Deletar Usuário",
                    use_container_width=True,
                    disabled=not confirm_delete,
                ):
                    if delete_user(selected_user):
                        st.session_state["user_action_message"] = f"✅ Usuário '{selected_user}' deletado!"
                        st.rerun()

            with col3:
                st.markdown("**Alterar Função**")
                user_data = next((u for u in users if u["username"] == selected_user), {})
                current_role = user_data.get("role", "user")
                all_roles = ["user", "admin", "supervisora", "catalog_editor"]
                role_idx = all_roles.index(current_role) if current_role in all_roles else 0
                new_role_edit = st.selectbox("Nova Função", all_roles, index=role_idx, key="edit_role_select")
                if st.button("✅ Atualizar Função", use_container_width=True):
                    db = get_firestore_client()
                    if db:
                        db.collection(FIRESTORE_USERS_COLLECTION).document(selected_user).update({"role": new_role_edit})
                        log_audit("update_user_role", {"username": selected_user, "new_role": new_role_edit})
                        st.session_state["user_action_message"] = f"✅ Função de '{selected_user}' alterada para '{new_role_edit}'!"
                        st.rerun()


def render_supervisora_mobile_view(user_info: Dict) -> None:
    """Mobile-first simplified interface for supervisoras."""
    st.title("📱 Novo Pedido")
    st.caption(f"Olá, {st.session_state.get('current_user')}! Crie seu pedido abaixo.")
    
    data = st.session_state.excel_data
    if data is None:
        st.error("⚠️ Dados não carregados. Contate o administrador.")
        return
    
    setores_df = data.get("Setor", pd.DataFrame())
    produtos_df = data.get("Produtos", pd.DataFrame())
    
    if setores_df.empty or produtos_df.empty:
        st.error("⚠️ Dados incompletos. Contate o administrador.")
        return
    
    # Filter setores for this supervisora
    setores_df = filter_setores_for_user(setores_df, user_info)
    
    if setores_df.empty:
        st.warning("⚠️ Nenhum setor vinculado a você. Contate o administrador.")
        return
    
    # Setor selection (full width, large)
    st.markdown("### 1️⃣ Selecione o Setor")
    setor_options = setores_df["label"].tolist()
    selected_setor_label = st.selectbox(
        "Setor/Cliente",
        options=setor_options,
        label_visibility="collapsed"
    )
    
    if selected_setor_label:
        selected_setor = setores_df[setores_df["label"] == selected_setor_label].iloc[0].to_dict()
        st.session_state.selected_setor = selected_setor
        
        st.divider()
        
        # Product list
        st.markdown("### 2️⃣ Adicione Produtos")
        
        # Initialize cart if needed
        if "cart" not in st.session_state:
            st.session_state.cart = []
        
        # Search box
        search = st.text_input("🔍 Buscar produto", "", key="search_mobile")
        
        # Filter products
        produtos_filtered = produtos_df.copy()
        if search:
            produtos_filtered = produtos_filtered[
                produtos_filtered["name"].str.contains(search, case=False, na=False) |
                produtos_filtered["productCode"].astype(str).str.contains(search, case=False, na=False)
            ]
        
        # Show products as cards (mobile-friendly)
        for _, produto in produtos_filtered.head(10).iterrows():
            with st.container(border=True):
                col1, col2 = st.columns([3, 1])
                with col1:
                    st.markdown(f"**{produto['name']}**")
                    st.caption(f"Código: {produto['productCode']} | R$ {produto['price']:.2f}")
                with col2:
                    qtde = st.number_input(
                        "Qtde",
                        min_value=0,
                        value=0,
                        key=f"qtde_mobile_{produto['productCode']}",
                        label_visibility="collapsed"
                    )
                    if qtde > 0:
                        if st.button("➕", key=f"add_mobile_{produto['productCode']}"):
                            add_item_to_cart(produto, int(qtde))
                            st.rerun()
        
        # Show cart
        if st.session_state.cart:
            st.divider()
            st.markdown("### 🛒 Carrinho")
            
            total = sum(item["total"] for item in st.session_state.cart)
            st.metric("Total", f"R$ {total:.2f}")
            
            for i, item in enumerate(st.session_state.cart):
                col1, col2, col3 = st.columns([3, 1, 1])
                with col1:
                    st.write(f"{item['nome']}")
                    st.caption(f"{item['qtde']}x R$ {item['preco']:.2f}")
                with col2:
                    st.write(f"R$ {item['total']:.2f}")
                with col3:
                    if st.button("🗑️", key=f"remove_cart_{i}"):
                        st.session_state.cart.pop(i)
                        st.rerun()
            
            # Save button
            st.divider()
            if st.button("✅ Enviar Pedido", use_container_width=True, type="primary"):
                persist_cart("Aguardando Aprovação")
                st.success("✅ Pedido enviado para aprovação!")
                st.balloons()
                st.rerun()


def main() -> None:
    st.set_page_config(page_title="Sistema de Gestão de Pedidos", layout="wide")
    init_session_state()

    # Check authentication
    if not check_authentication():
        render_login()
        st.stop()
    
    # Get user info
    current_user = st.session_state.get("current_user")
    user_info = get_user_info(current_user)
    
    # Supervisora gets mobile-first simplified interface
    if user_info["role"] == "supervisora":
        # Mobile header with logout
        col1, col2 = st.columns([3, 1])
        with col1:
            st.caption("📱 Sistema de Pedidos")
        with col2:
            if st.button("Sair", key="logout_mobile"):
                st.session_state.authenticated = False
                st.session_state.current_user = None
                st.rerun()
        
        render_sidebar()
        
        if st.session_state.excel_data is None:
            st.error("⚠️ Sistema não configurado. Contate o administrador.")
            st.stop()
        
        render_supervisora_mobile_view(user_info)
        st.stop()
    
    # Display current user in sidebar
    with st.sidebar:
        render_company_branding()
        st.divider()
        st.write(f"👤 Usuário: **{st.session_state.current_user}**")
        st.session_state.layout_mode_new = st.toggle(
            "Layout novo",
            value=bool(st.session_state.get("layout_mode_new", False)),
            help="Ativa o visual novo sem alterar dados do sistema.",
        )

        if st.session_state.get("layout_mode_new"):
            is_admin = st.session_state.get("current_user") == "admin"
            catalog_access = can_edit_catalog()
            nav_items = [
                ("novo", "🛒 Novo Pedido"),
                ("aguardando", "⏳ Aguardando Aprovação"),
                ("aprovados", "📋 Pedidos Aprovados"),
                ("historico", "📚 Histórico de Pedidos"),
            ]
            if is_admin:
                nav_items.append(("produtos", "📦 Atualizar Produtos"))
            if catalog_access:
                nav_items.append(("catalogo", "✏️ Editar Catálogo"))
            if is_admin:
                nav_items.append(("usuarios", "👥 Gerenciar Usuários"))

            if "layout_nav_page" not in st.session_state:
                st.session_state.layout_nav_page = "novo"

            valid_pages = {item[0] for item in nav_items}
            if st.session_state.layout_nav_page not in valid_pages:
                st.session_state.layout_nav_page = "novo"

            st.markdown("### 🧭 Navegação")
            for page_key, page_label in nav_items:
                btn_type = "primary" if st.session_state.layout_nav_page == page_key else "secondary"
                if st.button(
                    page_label,
                    key=f"sidebar_nav_{page_key}",
                    use_container_width=True,
                    type=btn_type,
                ):
                    st.session_state.layout_nav_page = page_key
                    st.rerun()

        if st.button("Sair"):
            st.session_state.authenticated = False
            st.session_state.current_user = None
            st.rerun()

    # Web interface header (after toggle state is resolved)
    if not st.session_state.get("layout_mode_new", False):
        st.title("Sistema de Gestão de Pedidos - Versão Web")
        st.caption("Interface Streamlit preparada para deploy em serviços como Railway ou Render.")

    apply_layout_theme(bool(st.session_state.get("layout_mode_new", False)))

    if st.session_state.get("layout_mode_new"):
        render_new_layout_header()
    else:
        st.caption("🧭 Modo de visualização ativo: Layout antigo")

    render_sidebar()

    if st.session_state.excel_data is None:
        st.stop()

    st.markdown('<div class="main-container">', unsafe_allow_html=True)

    if st.session_state.get("layout_mode_new"):
        current_page = st.session_state.get("layout_nav_page", "novo")
        if current_page == "novo":
            render_new_order_tab()
        elif current_page == "aguardando":
            render_aguardando_tab()
        elif current_page == "aprovados":
            render_approved_orders_tab()
        elif current_page == "historico":
            render_history_tab()
        elif current_page == "produtos" and st.session_state.get("current_user") == "admin":
            render_product_update_tab()
        elif current_page == "catalogo" and can_edit_catalog():
            render_catalog_editor_tab()
        elif current_page == "usuarios" and st.session_state.get("current_user") == "admin":
            render_user_management_tab()
        else:
            render_new_order_tab()
    else:
        # Create tabs based on user role (layout antigo)
        is_admin = st.session_state.get("current_user") == "admin"
        catalog_access = can_edit_catalog()
        if is_admin:
            tab_novo, tab_aguardando, tab_aprovados, tab_historico, tab_produtos, tab_catalogo, tab_usuarios = st.tabs([
                "Novo Pedido",
                "Aguardando Aprovação",
                "📋 Pedidos Aprovados",
                "📚 Histórico de Pedidos",
                "📦 Atualizar Produtos",
                "✏️ Editar Catálogo",
                "👥 Gerenciar Usuários"
            ])
        elif catalog_access:
            tab_novo, tab_aguardando, tab_aprovados, tab_historico, tab_catalogo = st.tabs([
                "Novo Pedido",
                "Aguardando Aprovação",
                "📋 Pedidos Aprovados",
                "📚 Histórico de Pedidos",
                "✏️ Editar Catálogo",
            ])
        else:
            tab_novo, tab_aguardando, tab_aprovados, tab_historico = st.tabs([
                "Novo Pedido",
                "Aguardando Aprovação",
                "📋 Pedidos Aprovados",
                "📚 Histórico de Pedidos"
            ])

        with tab_novo:
            render_new_order_tab()
        with tab_aguardando:
            render_aguardando_tab()
        with tab_aprovados:
            render_approved_orders_tab()
        with tab_historico:
            render_history_tab()

        if is_admin:
            with tab_produtos:
                render_product_update_tab()
            with tab_catalogo:
                render_catalog_editor_tab()
            with tab_usuarios:
                render_user_management_tab()
        elif catalog_access:
            with tab_catalogo:
                render_catalog_editor_tab()

    st.markdown('</div>', unsafe_allow_html=True)

    render_download_section()


if __name__ == "__main__":
    main()